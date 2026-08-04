import email
from sqlite3 import IntegrityError
from unittest import result
from django.db import models, transaction
from django.urls import reverse
from django.shortcuts import render, redirect
from django.contrib.auth import login, logout
from django.contrib.auth.decorators import login_required
from .forms import RegisterForm
from django.contrib import messages
from django.contrib.auth.models import User
from django.utils.crypto import get_random_string
from .models import AdminInviteCode, AdminRequest
from .forms import AdminRequestForm
from .forms import RegisterForm, JobRoleForm, SkillForm, JobRoleSkillForm
from .models import  JobRole, Skill, JobRoleSkill
from .forms import CareerMatchForm
from .models import CareerMatchResult
from .models import LearningResource
from .forms import LearningResourceForm
from django.shortcuts import  get_object_or_404
from .forms import UserProfileForm
from .models import UserProfile
from .utils import extract_text_from_resume, is_valid_resume, extract_skills_from_text

from .forms import ReadinessAssessmentForm
from .models import ReadinessAssessment
from .forms import IndustryToolForm, JobRoleToolForm

from .models import IndustryTool, JobRoleTool
from .forms import (
 
    ReadinessAssessmentForm
)
from django.db import transaction
from django.db.models import Avg
from .services.interview_evaluator import evaluate_answer
from django.core.paginator import Paginator
from django.db.models import Prefetch
from .models import CompetencyGroup, CompetencyGroupMember
from django.utils import timezone
from .forms import CompetencyGroupForm, CompetencyGroupMemberForm
from openpyxl import load_workbook
from .forms import DatasetImportForm
from .forms import BottleneckForm
from .models import EmployabilityBottleneck
from .models import CareerTransitionAnalysis
from .forms import CareerTransitionForm
from django.conf import settings
from .models import UserProject
from .forms import UserProjectForm
from django.contrib.auth.models import User
from .forms import InterviewSetupForm
from .models import InterviewSession
from .models import InterviewQuestion,InterviewAnswer
from .forms import InterviewAnswerForm
import logging

from django.db.models import Avg
from django.utils import timezone

from .services.interview_evaluator import evaluate_answer
from .services.ai_interview_evaluator import (
    AIInterviewEvaluationError,
    evaluate_answer_with_ai,
)
logger = logging.getLogger(__name__)
@login_required
def view_users(request):
    if not request.user.is_superuser:
        return redirect('user_dashboard')

    users = User.objects.filter(
        is_staff=False,
        is_superuser=False
    )

    return render(
        request,
        'career_app/view_users.html',
        {'users': users}
    )


@login_required
def view_admins(request):
    if not request.user.is_superuser:
        return redirect('user_dashboard')

    admins = User.objects.filter(
        is_staff=True,
        is_superuser=False
    )

    return render(
        request,
        'career_app/view_admins.html',
        {'admins': admins}
    )


def home(request):
    return render(request, 'career_app/home.html')


def dashboard_redirect(request):
    if request.user.is_superuser:
        return redirect('super_admin_dashboard')
    elif request.user.is_staff:
        return redirect('admin_dashboard')
    else:
        return redirect('user_dashboard')


def register(request):
    if request.method == 'POST':
        form = RegisterForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            return redirect('dashboard_redirect')
    else:
        form = RegisterForm()

    return render(request, 'career_app/register.html', {'form': form})


@login_required
def user_dashboard(request):
    return render(request, 'career_app/user_dashboard.html')


@login_required
def admin_dashboard(request):
    if not request.user.is_staff:
        return redirect('user_dashboard')
    return render(request, 'career_app/admin_dashboard.html')


@login_required
def super_admin_dashboard(request):
    if not request.user.is_superuser:
        return redirect('user_dashboard')
    return render(request, 'career_app/super_admin_dashboard.html')


def logout_view(request):
    logout(request)
    return redirect('home')

@login_required
def generate_admin_invite(request):
    if not request.user.is_superuser:
        return redirect('user_dashboard')

    code = None

    if request.method == 'POST':
        code = get_random_string(12).upper()
        AdminInviteCode.objects.create(
            code=code,
            created_by=request.user
        )

    return render(request, 'career_app/generate_admin_invite.html', {'code': code})


def admin_request(request):
    if request.method == 'POST':
        form = AdminRequestForm(request.POST)

        if form.is_valid():
            invite_code = form.cleaned_data['invite_code']

            try:
                code_obj = AdminInviteCode.objects.get(code=invite_code, is_used=False)
            except AdminInviteCode.DoesNotExist:
                messages.error(request, 'Invalid or already used invite code.')
                return redirect('admin_request')

            form.save()
            code_obj.is_used = True
            code_obj.save()

            messages.success(request, 'Admin request submitted successfully. Wait for Super Admin approval.')
            return redirect('login')
    else:
        form = AdminRequestForm()

    return render(request, 'career_app/admin_request.html', {'form': form})


@login_required
def view_admin_requests(request):
    if not request.user.is_superuser:
        return redirect('user_dashboard')

    requests = AdminRequest.objects.all().order_by('-created_at')
    return render(request, 'career_app/view_admin_requests.html', {'requests': requests})


@login_required
def approve_admin_request(request, request_id):
    if not request.user.is_superuser:
        return redirect('user_dashboard')

    admin_req = AdminRequest.objects.get(id=request_id)

    if User.objects.filter(email=admin_req.email).exists():
        messages.error(request, 'A user with this email already exists.')
        return redirect('view_admin_requests')

    username_base = admin_req.email.split('@')[0]
    username = username_base
    counter = 1

    while User.objects.filter(username=username).exists():
        username = f"{username_base}{counter}"
        counter += 1

    temp_password = get_random_string(10)

    User.objects.create_user(
        username=username,
        email=admin_req.email,
        password=temp_password,
        first_name=admin_req.full_name,
        is_staff=True
    )

    admin_req.status = 'Approved'
    admin_req.save()

    return render(request, 'career_app/admin_created.html', {
        'username': username,
        'temp_password': temp_password
    })


@login_required
def reject_admin_request(request, request_id):
    if not request.user.is_superuser:
        return redirect('user_dashboard')

    admin_req = AdminRequest.objects.get(id=request_id)
    admin_req.status = 'Rejected'
    admin_req.save()

    return redirect('view_admin_requests')
from django.core.mail import send_mail
from django.http import HttpResponse
@login_required
def add_job_role(request):
    if not request.user.is_staff:
        return redirect('user_dashboard')

    form = JobRoleForm(request.POST or None)

    if form.is_valid():
        form.save()
        return redirect('view_job_roles')

    return render(request, 'career_app/add_job_role.html', {'form': form})


@login_required
def view_job_roles(request):
    if not request.user.is_staff:
        return redirect('user_dashboard')

    roles = JobRole.objects.all()
    return render(request, 'career_app/view_job_roles.html', {'roles': roles})

@login_required
def edit_job_role(request, role_id):
    if not request.user.is_staff:
        return redirect('user_dashboard')

    role = JobRole.objects.get(id=role_id)

    form = JobRoleForm(
        request.POST or None,
        instance=role
    )

    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, "Job role updated successfully.")
        return redirect('view_job_roles')

    return render(
        request,
        'career_app/edit_job_role.html',
        {
            'form': form,
            'role': role
        }
    )


@login_required
def delete_job_role(request, role_id):
    if not request.user.is_staff:
        return redirect('user_dashboard')

    role = JobRole.objects.get(id=role_id)

    if request.method == 'POST':
        role.delete()
        messages.success(request, "Job role deleted successfully.")
        return redirect('view_job_roles')

    return render(
        request,
        'career_app/delete_job_role.html',
        {
            'role': role
        }
    )

@login_required
def add_skill(request):
    if not request.user.is_staff:
        return redirect('user_dashboard')

    form = SkillForm(request.POST or None)

    if form.is_valid():
        form.save()
        return redirect('view_skills')

    return render(request, 'career_app/add_skill.html', {'form': form})


@login_required
def view_skills(request):
    if not request.user.is_staff:
        return redirect('user_dashboard')

    skills = Skill.objects.all()
    return render(request, 'career_app/view_skills.html', {'skills': skills})

@login_required
def edit_skill(request, skill_id):
    if not request.user.is_staff:
        return redirect('user_dashboard')

    skill = Skill.objects.get(id=skill_id)

    form = SkillForm(
        request.POST or None,
        instance=skill
    )

    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, "Skill updated successfully.")
        return redirect('view_skills')

    return render(
        request,
        'career_app/edit_skill.html',
        {
            'form': form,
            'skill': skill
        }
    )


@login_required
def delete_skill(request, skill_id):
    if not request.user.is_staff:
        return redirect('user_dashboard')

    skill = Skill.objects.get(id=skill_id)

    if request.method == 'POST':
        skill.delete()
        messages.success(request, "Skill deleted successfully.")
        return redirect('view_skills')

    return render(
        request,
        'career_app/delete_skill.html',
        {
            'skill': skill
        }
    )
@login_required
def assign_skill_to_role(request):
    if not request.user.is_staff:
        return redirect('user_dashboard')

    form = JobRoleSkillForm(request.POST or None)

    if request.method == 'POST' and form.is_valid():
        job_role = form.cleaned_data['job_role']

        importance_groups = {
            'High': form.cleaned_data['high_skills'],
            'Medium': form.cleaned_data['medium_skills'],
            'Low': form.cleaned_data['low_skills'],
        }

        for importance, skills in importance_groups.items():
            for skill in skills:
                obj, created = JobRoleSkill.objects.get_or_create(
                    job_role=job_role,
                    skill=skill,
                    defaults={'importance': importance}
                )

                if not created:
                    obj.importance = importance
                    obj.save()

        return redirect('view_role_skills')

    return render(request, 'career_app/assign_skill_to_role.html', {'form': form})


@login_required
def view_role_skills(request):
    if not request.user.is_staff:
        return redirect('user_dashboard')

    role_skills = JobRoleSkill.objects.select_related('job_role', 'skill').all()
    return render(request, 'career_app/view_role_skills.html', {'role_skills': role_skills})

@login_required
def career_match(request):
    form = CareerMatchForm(request.POST or None)

    if request.method == 'POST' and form.is_valid():
        profile = UserProfile.objects.filter(
            user=request.user
        ).first()

        if not profile:
            return redirect('create_profile')

        result = form.save(commit=False)
        result.user = request.user
        result.save()

        user_skills = (
            profile.extracted_skills.all()
            | profile.manual_skills.all()
        ).distinct()

        result.selected_skills.set(user_skills)

        user_skill_ids = set(
            user_skills.values_list('id', flat=True)
        )

        required_skills = Skill.objects.filter(
            jobroleskill__job_role=result.job_role
        ).distinct()

        if not required_skills.exists():
            result.match_score = 0
            result.save()

            result.matched_skills.clear()
            result.missing_skills.clear()

            return redirect(
                'career_match_result',
                result_id=result.id
            )

        matched_skill_ids = set()
        missing_skill_ids = set()
        grouped_skill_ids = set()

        satisfied_requirement_count = 0
        total_requirement_count = 0

        groups = CompetencyGroup.objects.filter(
            job_role=result.job_role
        ).prefetch_related(
            Prefetch(
                'members',
                queryset=CompetencyGroupMember.objects.select_related(
                    'job_role_skill__skill'
                )
            )
        )

        for group in groups:
            members = list(group.members.all())

            skill_ids = {
                member.job_role_skill.skill_id
                for member in members
            }

            if not skill_ids:
                continue

            grouped_skill_ids.update(skill_ids)

            matched_ids = skill_ids.intersection(
                user_skill_ids
            )

            if group.rule == 'ANY_ONE':
                # The entire group counts as one requirement.
                total_requirement_count += 1

                if matched_ids:
                    satisfied_requirement_count += 1

                    # Store the actual option(s) the student knows.
                    matched_skill_ids.update(matched_ids)

                # If nothing is matched, do not add every option
                # to missing_skills. The result page will display
                # this as one missing competency group.

            elif group.rule == 'ALL_REQUIRED':
                # Every member is an individual requirement.
                total_requirement_count += len(skill_ids)
                satisfied_requirement_count += len(matched_ids)

                matched_skill_ids.update(matched_ids)

                missing_skill_ids.update(
                    skill_ids.difference(user_skill_ids)
                )

        standalone_skills = required_skills.exclude(
            id__in=grouped_skill_ids
        )

        for skill in standalone_skills:
            total_requirement_count += 1

            if skill.id in user_skill_ids:
                satisfied_requirement_count += 1
                matched_skill_ids.add(skill.id)
            else:
                missing_skill_ids.add(skill.id)

        if total_requirement_count > 0:
            match_score = (
                satisfied_requirement_count
                / total_requirement_count
            ) * 100
        else:
            match_score = 0

        result.match_score = round(match_score, 2)
        result.save()

        result.matched_skills.set(
            Skill.objects.filter(
                id__in=matched_skill_ids
            )
        )

        result.missing_skills.set(
            Skill.objects.filter(
                id__in=missing_skill_ids
            )
        )

        return redirect(
            'career_match_result',
            result_id=result.id
        )

    return render(
        request,
        'career_app/career_match.html',
        {
            'form': form
        }
    )
@login_required
def career_match_result(request, result_id):
    result = CareerMatchResult.objects.get(
        id=result_id,
        user=request.user
    )

    user_skill_ids = set(
        result.selected_skills.values_list(
            'id',
            flat=True
        )
    )

    grouped_skill_ids = set()
    competency_results = []

    groups = CompetencyGroup.objects.filter(
        job_role=result.job_role
    ).prefetch_related(
        Prefetch(
            'members',
            queryset=CompetencyGroupMember.objects.select_related(
                'job_role_skill__skill'
            ).order_by(
                'job_role_skill__skill__skill_name'
            )
        )
    ).order_by(
        'group_name'
    )

    for group in groups:
        members = list(
            group.members.all()
        )

        skills = [
            member.job_role_skill.skill
            for member in members
        ]

        skill_ids = {
            skill.id
            for skill in skills
        }

        if not skill_ids:
            continue

        grouped_skill_ids.update(
            skill_ids
        )

        matched_skills = [
            skill
            for skill in skills
            if skill.id in user_skill_ids
        ]

        missing_skills = [
            skill
            for skill in skills
            if skill.id not in user_skill_ids
        ]

        if group.rule == 'ANY_ONE':
            is_satisfied = bool(
                matched_skills
            )

            competency_results.append({
                'group_name': group.group_name,
                'rule': group.rule,
                'rule_label': 'Choose any one',
                'status': (
                    'SATISFIED'
                    if is_satisfied
                    else 'MISSING'
                ),
                'is_satisfied': is_satisfied,
                'is_partial': False,
                'matched_skills': matched_skills,
                'missing_skills': [],
                'options': skills,
            })

        else:
            is_satisfied = (
                len(missing_skills) == 0
            )

            is_partial = (
                bool(matched_skills)
                and bool(missing_skills)
            )

            competency_results.append({
                'group_name': group.group_name,
                'rule': group.rule,
                'rule_label': 'All required',
                'status': (
                    'SATISFIED'
                    if is_satisfied
                    else (
                        'PARTIAL'
                        if is_partial
                        else 'MISSING'
                    )
                ),
                'is_satisfied': is_satisfied,
                'is_partial': is_partial,
                'matched_skills': matched_skills,
                'missing_skills': missing_skills,
                'options': skills,
            })

    standalone_skills = Skill.objects.filter(
        jobroleskill__job_role=result.job_role
    ).exclude(
        id__in=grouped_skill_ids
    ).distinct().order_by(
        'skill_name'
    )

    standalone_results = []

    for skill in standalone_skills:
        is_satisfied = (
            skill.id in user_skill_ids
        )

        standalone_results.append({
            'skill': skill,
            'is_satisfied': is_satisfied,
        })

    missing_all_required_skill_ids = set()

    for competency in competency_results:
        if competency['rule'] == 'ALL_REQUIRED':
            missing_all_required_skill_ids.update(
                skill.id
                for skill in competency['missing_skills']
            )

    missing_standalone_skill_ids = {
        item['skill'].id
        for item in standalone_results
        if not item['is_satisfied']
    }

    learning_skill_ids = (
        missing_all_required_skill_ids
        | missing_standalone_skill_ids
    )

    resources = LearningResource.objects.filter(
        skill_id__in=learning_skill_ids
    ).select_related(
        'skill'
    ).order_by(
        'skill__skill_name',
        'title'
    )

    satisfied_competencies = sum(
        1
        for item in competency_results
        if item['is_satisfied']
    )

    total_competencies = len(
        competency_results
    )

    return render(
        request,
        'career_app/career_match_result.html',
        {
            'result': result,
            'competency_results': competency_results,
            'standalone_results': standalone_results,
            'resources': resources,
            'satisfied_competencies': satisfied_competencies,
            'total_competencies': total_competencies,
        }
    )
@login_required
def add_learning_resource(request):
    if not request.user.is_staff:
        return redirect('user_dashboard')

    form = LearningResourceForm(request.POST or None)

    if form.is_valid():
        form.save()
        return redirect('view_learning_resources')

    return render(
        request,
        'career_app/add_learning_resource.html',
        {'form': form}
    )


@login_required
def view_learning_resources(request):
    if not request.user.is_staff:
        return redirect('user_dashboard')

    resources = LearningResource.objects.select_related('skill')

    return render(
        request,
        'career_app/view_learning_resources.html',
        {'resources': resources}
    )

@login_required
def edit_learning_resource(request, resource_id):
    if not request.user.is_staff:
        return redirect('user_dashboard')

    resource = LearningResource.objects.get(id=resource_id)

    form = LearningResourceForm(
        request.POST or None,
        instance=resource
    )

    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, "Learning resource updated successfully.")
        return redirect('view_learning_resources')

    return render(
        request,
        'career_app/edit_learning_resource.html',
        {
            'form': form,
            'resource': resource
        }
    )


@login_required
def delete_learning_resource(request, resource_id):
    if not request.user.is_staff:
        return redirect('user_dashboard')

    resource = LearningResource.objects.get(id=resource_id)

    if request.method == 'POST':
        resource.delete()
        messages.success(request, "Learning resource deleted successfully.")
        return redirect('view_learning_resources')

    return render(
        request,
        'career_app/delete_learning_resource.html',
        {
            'resource': resource
        }
    )
@login_required
def learning_roadmap(request, result_id):

    result = CareerMatchResult.objects.get(
        id=result_id,
        user=request.user
    )

    missing_skills = result.missing_skills.all()

    roadmap_steps = []

    step_number = 1

    for skill in missing_skills:
        roadmap_steps.append(
            f"Step {step_number}: Learn {skill.skill_name}"
        )
        step_number += 1

    roadmap_steps.append(
        f"Step {step_number}: Build a project related to {result.job_role.role_name}"
    )

    return render(
        request,
        'career_app/learning_roadmap.html',
        {
            'result': result,
            'roadmap_steps': roadmap_steps
        }
    )
@login_required
def create_profile(request):
    if UserProfile.objects.filter(user=request.user).exists():
        return redirect('view_profile')

    form = UserProfileForm(request.POST or None, request.FILES or None)

    if form.is_valid():
        profile = form.save(commit=False)
        profile.user = request.user
        profile.save()
        form.save_m2m()

        if profile.resume:
            file_path = profile.resume.path

            extracted_text = extract_text_from_resume(file_path)
            profile.extracted_text = extracted_text
            profile.is_resume_valid = is_valid_resume(extracted_text)
            profile.save()

            if profile.is_resume_valid:
                all_skills = Skill.objects.all()
                extracted_skills = extract_skills_from_text(
                    extracted_text,
                    all_skills
                )
                profile.extracted_skills.set(extracted_skills)

        return redirect('view_profile')

    return render(request, 'career_app/create_profile.html', {'form': form})


@login_required
def view_profile(request):
    profile = UserProfile.objects.filter(user=request.user).first()
    return render(request, 'career_app/view_profile.html', {'profile': profile})


@login_required
def edit_profile(request):
    profile = UserProfile.objects.get(user=request.user)

    form = UserProfileForm(
        request.POST or None,
        request.FILES or None,
        instance=profile
    )

    if form.is_valid():
        profile = form.save()

        if profile.resume:
            file_path = profile.resume.path

            extracted_text = extract_text_from_resume(file_path)
            profile.extracted_text = extracted_text
            profile.is_resume_valid = is_valid_resume(extracted_text)
            profile.save()

            if profile.is_resume_valid:
                all_skills = Skill.objects.all()
                extracted_skills = extract_skills_from_text(
                    extracted_text,
                    all_skills
                )
                profile.extracted_skills.set(extracted_skills)
            else:
                profile.extracted_skills.clear()

        return redirect('view_profile')

    return render(request, 'career_app/edit_profile.html', {'form': form})

@login_required
def readiness_assessment(request):
    form = ReadinessAssessmentForm(
        request.POST or None
    )

    if request.method == 'POST' and form.is_valid():
        assessment = form.save(
            commit=False
        )

        assessment.user = request.user
        job_role = assessment.job_role

        profile = UserProfile.objects.filter(
            user=request.user
        ).first()

        if profile:
            user_skills = (
                profile.extracted_skills.all()
                | profile.manual_skills.all()
            ).distinct()

            user_tools = profile.manual_tools.all()

        else:
            user_skills = Skill.objects.none()
            user_tools = IndustryTool.objects.none()

        user_skill_ids = set(
            user_skills.values_list(
                'id',
                flat=True
            )
        )

        user_tool_ids = set(
            user_tools.values_list(
                'id',
                flat=True
            )
        )

        required_skills = Skill.objects.filter(
            jobroleskill__job_role=job_role
        ).distinct()

        required_tools = IndustryTool.objects.filter(
            jobroletool__job_role=job_role
        ).distinct()

        matched_skill_ids = set()
        missing_skill_ids = set()
        grouped_skill_ids = set()

        missing_any_one_groups = []
        satisfied_any_one_groups = []

        satisfied_requirement_count = 0
        total_requirement_count = 0

        groups = CompetencyGroup.objects.filter(
            job_role=job_role
        ).order_by(
            'group_name'
        )

        for group in groups:
            members = CompetencyGroupMember.objects.filter(
                group=group
            ).select_related(
                'job_role_skill__skill'
            ).order_by(
                'job_role_skill__skill__skill_name'
            )

            skills = [
                member.job_role_skill.skill
                for member in members
            ]

            skill_ids = {
                skill.id
                for skill in skills
            }

            if not skill_ids:
                continue

            grouped_skill_ids.update(
                skill_ids
            )

            matched_ids = skill_ids.intersection(
                user_skill_ids
            )

            if group.rule == 'ANY_ONE':
                # The whole group counts as one requirement.
                total_requirement_count += 1

                if matched_ids:
                    satisfied_requirement_count += 1

                    matched_skill_ids.update(
                        matched_ids
                    )

                    matched_names = [
                        skill.skill_name
                        for skill in skills
                        if skill.id in matched_ids
                    ]

                    satisfied_any_one_groups.append({
                        'group_name': group.group_name,
                        'matched_skills': matched_names,
                    })

                else:
                    option_names = [
                        skill.skill_name
                        for skill in skills
                    ]

                    missing_any_one_groups.append({
                        'group_name': group.group_name,
                        'options': option_names,
                    })

            elif group.rule == 'ALL_REQUIRED':
                # Every member counts as an individual requirement.
                total_requirement_count += len(
                    skill_ids
                )

                satisfied_requirement_count += len(
                    matched_ids
                )

                matched_skill_ids.update(
                    matched_ids
                )

                missing_skill_ids.update(
                    skill_ids.difference(
                        user_skill_ids
                    )
                )

        standalone_skills = required_skills.exclude(
            id__in=grouped_skill_ids
        )

        for skill in standalone_skills:
            total_requirement_count += 1

            if skill.id in user_skill_ids:
                satisfied_requirement_count += 1
                matched_skill_ids.add(
                    skill.id
                )
            else:
                missing_skill_ids.add(
                    skill.id
                )

        matched_skills = Skill.objects.filter(
            id__in=matched_skill_ids
        ).order_by(
            'skill_name'
        )

        missing_skills = Skill.objects.filter(
            id__in=missing_skill_ids
        ).order_by(
            'skill_name'
        )

        if total_requirement_count > 0:
            academic_score = (
                satisfied_requirement_count
                / total_requirement_count
            ) * 100
        else:
            academic_score = 0

        required_tool_ids = set(
            required_tools.values_list(
                'id',
                flat=True
            )
        )

        matched_tool_ids = required_tool_ids.intersection(
            user_tool_ids
        )

        missing_tool_ids = required_tool_ids.difference(
            user_tool_ids
        )

        matched_tools = IndustryTool.objects.filter(
            id__in=matched_tool_ids
        ).order_by(
            'tool_name'
        )

        missing_tools = IndustryTool.objects.filter(
            id__in=missing_tool_ids
        ).order_by(
            'tool_name'
        )

        if required_tool_ids:
            industry_score = (
                len(matched_tool_ids)
                / len(required_tool_ids)
            ) * 100
        else:
            industry_score = 0

        overall_score = (
            academic_score * 0.6
        ) + (
            industry_score * 0.4
        )

        assessment.academic_score = round(
            academic_score,
            2
        )

        assessment.industry_score = round(
            industry_score,
            2
        )

        assessment.overall_readiness_score = round(
            overall_score,
            2
        )

        matched_skill_names = list(
            matched_skills.values_list(
                'skill_name',
                flat=True
            )
        )

        missing_skill_names = list(
            missing_skills.values_list(
                'skill_name',
                flat=True
            )
        )

        matched_tool_names = list(
            matched_tools.values_list(
                'tool_name',
                flat=True
            )
        )

        missing_tool_names = list(
            missing_tools.values_list(
                'tool_name',
                flat=True
            )
        )

        strength_parts = []

        if matched_skill_names:
            strength_parts.append(
                "Matched Skills: "
                + ", ".join(
                    matched_skill_names
                )
            )
        else:
            strength_parts.append(
                "Matched Skills: None"
            )

        for group_data in satisfied_any_one_groups:
            matched_options = ", ".join(
                group_data['matched_skills']
            )

            strength_parts.append(
                f"{group_data['group_name']}: "
                f"satisfied with {matched_options}"
            )

        if matched_tool_names:
            strength_parts.append(
                "Matched Tools: "
                + ", ".join(
                    matched_tool_names
                )
            )
        else:
            strength_parts.append(
                "Matched Tools: None"
            )

        assessment.strengths = "\n\n".join(
            strength_parts
        )

        weakness_parts = []

        if missing_skill_names:
            weakness_parts.append(
                "Missing Required Skills: "
                + ", ".join(
                    missing_skill_names
                )
            )

        for group_data in missing_any_one_groups:
            options = ", ".join(
                group_data['options']
            )

            weakness_parts.append(
                f"{group_data['group_name']}: "
                f"choose any one of {options}"
            )

        if missing_tool_names:
            weakness_parts.append(
                "Missing Tools: "
                + ", ".join(
                    missing_tool_names
                )
            )

        if not weakness_parts:
            weakness_parts.append(
                "No major skill or tool gaps found."
            )

        assessment.weaknesses = "\n\n".join(
            weakness_parts
        )

        if overall_score >= 75:
            assessment.recommendation = (
                "You are close to industry-ready for this role. "
                "Focus on portfolio projects and interview preparation."
            )

        elif overall_score >= 50:
            assessment.recommendation = (
                "You have a moderate readiness level. "
                "Complete the missing competencies, improve required "
                "tools, and build practical projects."
            )

        else:
            assessment.recommendation = (
                "Your readiness is low for this role. "
                "Start with the missing core competencies and required "
                "industry tools before applying."
            )

        assessment.save()

        return redirect(
            'readiness_result',
            assessment_id=assessment.id
        )

    return render(
        request,
        'career_app/readiness_assessment.html',
        {
            'form': form
        }
    )
@login_required
def readiness_result(request, assessment_id):
    assessment = ReadinessAssessment.objects.get(
        id=assessment_id,
        user=request.user
    )

    return render(
        request,
        'career_app/readiness_result.html',
        {'assessment': assessment}
    )
@login_required
def add_industry_tool(request):
    if not request.user.is_staff:
        return redirect('user_dashboard')

    form = IndustryToolForm(request.POST or None)

    if form.is_valid():
        form.save()
        return redirect('view_industry_tools')

    return render(request, 'career_app/add_industry_tool.html', {'form': form})


@login_required
def view_industry_tools(request):
    if not request.user.is_staff:
        return redirect('user_dashboard')

    tools = IndustryTool.objects.all()
    return render(request, 'career_app/view_industry_tools.html', {'tools': tools})

@login_required
def edit_industry_tool(request, tool_id):
    if not request.user.is_staff:
        return redirect('user_dashboard')

    tool = IndustryTool.objects.get(id=tool_id)

    form = IndustryToolForm(
        request.POST or None,
        instance=tool
    )

    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, "Industry tool updated successfully.")
        return redirect('view_industry_tools')

    return render(
        request,
        'career_app/edit_industry_tool.html',
        {
            'form': form,
            'tool': tool
        }
    )


@login_required
def delete_industry_tool(request, tool_id):
    if not request.user.is_staff:
        return redirect('user_dashboard')

    tool = IndustryTool.objects.get(id=tool_id)

    if request.method == 'POST':
        tool.delete()
        messages.success(request, "Industry tool deleted successfully.")
        return redirect('view_industry_tools')

    return render(
        request,
        'career_app/delete_industry_tool.html',
        {
            'tool': tool
        }
    )

@login_required
def assign_tool_to_role(request):
    if not request.user.is_staff:
        return redirect('user_dashboard')

    form = JobRoleToolForm(request.POST or None)

    if request.method == 'POST' and form.is_valid():
        job_role = form.cleaned_data['job_role']

        importance_groups = {
            'High': form.cleaned_data['high_tools'],
            'Medium': form.cleaned_data['medium_tools'],
            'Low': form.cleaned_data['low_tools'],
        }

        for importance, tools in importance_groups.items():
            for tool in tools:
                obj, created = JobRoleTool.objects.get_or_create(
                    job_role=job_role,
                    tool=tool,
                    defaults={'importance': importance}
                )

                if not created:
                    obj.importance = importance
                    obj.save()

        return redirect('view_role_tools')

    return render(request, 'career_app/assign_tool_to_role.html', {'form': form})


@login_required
def view_role_tools(request):
    if not request.user.is_staff:
        return redirect('user_dashboard')

    role_tools = JobRoleTool.objects.select_related('job_role', 'tool').all()
    return render(request, 'career_app/view_role_tools.html', {'role_tools': role_tools})


@login_required
def bottleneck_detection(request):
    form = BottleneckForm(request.POST or None)

    if request.method == 'POST':
        if form.is_valid():
            job_role = form.cleaned_data['job_role']

            assessment = ReadinessAssessment.objects.filter(
                user=request.user,
                job_role=job_role
            ).order_by('-created_at').first()

            if not assessment:
                return redirect('readiness_assessment')

            projects = UserProject.objects.filter(user=request.user)
            project_count = projects.count()

            required_skills = Skill.objects.filter(
                jobroleskill__job_role=job_role
            ).distinct()

            required_tools = IndustryTool.objects.filter(
                jobroletool__job_role=job_role
            ).distinct()

            project_skill_ids = set()
            project_tool_ids = set()

            for project in projects:
                project_skill_ids.update(
                    project.skills_used.values_list('id', flat=True)
                )
                project_tool_ids.update(
                    project.tools_used.values_list('id', flat=True)
                )

            matched_project_skill_ids = set()
            missing_project_skill_ids = set()
            grouped_skill_ids = set()

            groups = CompetencyGroup.objects.filter(
                job_role=job_role
            )

            for group in groups:
                members = CompetencyGroupMember.objects.filter(
                    group=group
                ).select_related('job_role_skill__skill')

                skills = [
                    member.job_role_skill.skill
                    for member in members
                ]

                skill_ids = {
                    skill.id
                    for skill in skills
                }

                grouped_skill_ids.update(skill_ids)

                if group.rule == "ANY_ONE":
                    matched = skill_ids.intersection(project_skill_ids)

                    if matched:
                        matched_project_skill_ids.add(next(iter(matched)))
                    else:
                        missing_project_skill_ids.update(skill_ids)

                else:
                    for skill in skills:
                        if skill.id in project_skill_ids:
                            matched_project_skill_ids.add(skill.id)
                        else:
                            missing_project_skill_ids.add(skill.id)

            normal_required_skills = required_skills.exclude(
                id__in=grouped_skill_ids
            )

            for skill in normal_required_skills:
                if skill.id in project_skill_ids:
                    matched_project_skill_ids.add(skill.id)
                else:
                    missing_project_skill_ids.add(skill.id)

            required_tool_ids = set(
                required_tools.values_list('id', flat=True)
            )

            relevant_project_tools = project_tool_ids.intersection(
                required_tool_ids
            )

            missing_project_tool_ids = required_tool_ids.difference(
                project_tool_ids
            )

            total_skill_items = (
                len(matched_project_skill_ids) +
                len(missing_project_skill_ids)
            )

            if total_skill_items > 0:
                project_skill_coverage = round(
                    (len(matched_project_skill_ids) / total_skill_items) * 100,
                    2
                )
            else:
                project_skill_coverage = 0

            if len(required_tool_ids) > 0:
                project_tool_coverage = round(
                    (len(relevant_project_tools) / len(required_tool_ids)) * 100,
                    2
                )
            else:
                project_tool_coverage = 0

            total_required_items = total_skill_items + len(required_tool_ids)
            total_relevant_items = (
                len(matched_project_skill_ids) +
                len(relevant_project_tools)
            )

            if total_required_items > 0:
                project_relevance_score = round(
                    (total_relevant_items / total_required_items) * 100,
                    2
                )
            else:
                project_relevance_score = 0

            missing_project_skills = Skill.objects.filter(
                id__in=missing_project_skill_ids
            )

            missing_project_tools = IndustryTool.objects.filter(
                id__in=missing_project_tool_ids
            )

            missing_skill_names = ", ".join(
                missing_project_skills.values_list("skill_name", flat=True)
            )

            missing_tool_names = ", ".join(
                missing_project_tools.values_list("tool_name", flat=True)
            )

            bottleneck = EmployabilityBottleneck()
            bottleneck.user = request.user
            bottleneck.job_role = job_role
            bottleneck.readiness_assessment = assessment

            if assessment.academic_score < 50:
                bottleneck.main_bottleneck = "Skill Deficiency"
                bottleneck.explanation = (
                    f"Your academic readiness score is {assessment.academic_score}%, "
                    f"which means your core skill foundation for {job_role.role_name} is weak."
                )
                bottleneck.recommendation = (
                    "First improve the missing core skills for this role before focusing on projects."
                )

            elif assessment.industry_score < 50:
                bottleneck.main_bottleneck = "Industry Tool Deficiency"
                bottleneck.explanation = (
                    f"Your industry tool readiness score is {assessment.industry_score}%, "
                    f"which means you are missing important tools required for {job_role.role_name}."
                )
                bottleneck.recommendation = (
                    "Learn the missing tools and use them inside practical projects."
                )

            elif project_count == 0:
                bottleneck.main_bottleneck = "Lack of Practical Projects"
                bottleneck.explanation = (
                    "You have not added any projects to prove practical experience."
                )
                bottleneck.recommendation = (
                    f"Add at least one {job_role.role_name} project and map the skills/tools used."
                )

            elif project_skill_coverage < 50:
                bottleneck.main_bottleneck = "Weak Project Skill Evidence"
                bottleneck.explanation = (
                    f"You added {project_count} project(s), but they only prove "
                    f"{project_skill_coverage}% of the required skill competencies for {job_role.role_name}."
                )
                bottleneck.recommendation = (
                    f"Strengthen your projects using missing role skills such as: "
                    f"{missing_skill_names if missing_skill_names else 'more target-role skills'}."
                )

            elif project_tool_coverage < 50:
                bottleneck.main_bottleneck = "Weak Project Tool Evidence"
                bottleneck.explanation = (
                    f"You added {project_count} project(s), but they only show "
                    f"{project_tool_coverage}% of the required tools for {job_role.role_name}."
                )
                bottleneck.recommendation = (
                    f"Update your projects to include tools such as: "
                    f"{missing_tool_names if missing_tool_names else 'more industry tools'}."
                )

            elif project_relevance_score < 70:
                bottleneck.main_bottleneck = "Weak Project Relevance"
                bottleneck.explanation = (
                    f"You added {project_count} project(s), but their combined relevance score is only "
                    f"{project_relevance_score}% for {job_role.role_name}."
                )
                bottleneck.recommendation = (
                    "Build one stronger role-specific project instead of adding many weak or unrelated projects."
                )

            elif assessment.overall_readiness_score < 75:
                bottleneck.main_bottleneck = "Moderate Readiness"
                bottleneck.explanation = (
                    f"Your projects are relevant, but your overall readiness score is "
                    f"{assessment.overall_readiness_score}%, which is still below industry-ready level."
                )
                bottleneck.recommendation = (
                    "Improve your weakest readiness area and polish your best project into portfolio quality."
                )

            else:
                bottleneck.main_bottleneck = "No Major Bottleneck"
                bottleneck.explanation = (
                    f"You appear ready for {job_role.role_name} based on your current skills, tools, and projects."
                )
                bottleneck.recommendation = (
                    "Focus on interview preparation, portfolio polishing, and job applications."
                )

            bottleneck.save()

            return redirect(
                'bottleneck_result',
                bottleneck_id=bottleneck.id
            )

    return render(
        request,
        'career_app/bottleneck_detection.html',
        {'form': form}
    )
@login_required
def bottleneck_result(request, bottleneck_id):
    bottleneck = EmployabilityBottleneck.objects.get(
        id=bottleneck_id,
        user=request.user
    )

    return render(
        request,
        'career_app/bottleneck_result.html',
        {'bottleneck': bottleneck}
    )

@login_required
def add_project(request):

    form = UserProjectForm(request.POST or None)

    if request.method == 'POST':
        if form.is_valid():

            project = form.save(commit=False)
            project.user = request.user
            project.save()

            form.save_m2m()

            return redirect('view_projects')

    return render(
        request,
        'career_app/add_project.html',
        {'form': form}
    )

@login_required
def view_projects(request):

    projects = UserProject.objects.filter(
        user=request.user
    )

    return render(
        request,
        'career_app/view_projects.html',
        {'projects': projects}
    )

@login_required
def delete_project(request, project_id):
    project = UserProject.objects.get(
        id=project_id,
        user=request.user
    )

    project.delete()

    return redirect('view_projects')

@login_required
def career_transition_analysis(request):
    form = CareerTransitionForm(request.POST or None)

    if request.method == 'POST' and form.is_valid():
        current_role = form.cleaned_data['current_role']
        target_role = form.cleaned_data['target_role']

        if current_role == target_role:
            messages.error(request, "Current role and target role cannot be the same.")
            return redirect('career_transition_analysis')

        profile = UserProfile.objects.filter(user=request.user).first()

        if not profile:
            return redirect('create_profile')

        user_skills = (
            profile.extracted_skills.all() | profile.manual_skills.all()
        ).distinct()

        user_tools = profile.manual_tools.all()

        target_required_skills = Skill.objects.filter(
            jobroleskill__job_role=target_role
        ).distinct()

        target_required_tools = IndustryTool.objects.filter(
            jobroletool__job_role=target_role
        ).distinct()

        matched_skill_ids = set()
        missing_skill_ids = set()

        user_skill_ids = set(
            user_skills.values_list('id', flat=True)
        )

        grouped_skill_ids = set()

        groups = CompetencyGroup.objects.filter(
            job_role=target_role
        )

        for group in groups:
            members = CompetencyGroupMember.objects.filter(
                group=group
            ).select_related('job_role_skill__skill')

            skills = [
                member.job_role_skill.skill
                for member in members
            ]

            skill_ids = {
                skill.id
                for skill in skills
            }

            grouped_skill_ids.update(skill_ids)

            if group.rule == "ANY_ONE":
                matched = skill_ids.intersection(user_skill_ids)

                if matched:
                    matched_skill_ids.add(next(iter(matched)))
                else:
                    missing_skill_ids.update(skill_ids)

            else:
                for skill in skills:
                    if skill.id in user_skill_ids:
                        matched_skill_ids.add(skill.id)
                    else:
                        missing_skill_ids.add(skill.id)

        normal_skills = target_required_skills.exclude(
            id__in=grouped_skill_ids
        )

        for skill in normal_skills:
            if skill.id in user_skill_ids:
                matched_skill_ids.add(skill.id)
            else:
                missing_skill_ids.add(skill.id)

        matched_skills = Skill.objects.filter(
            id__in=matched_skill_ids
        )

        missing_skills = Skill.objects.filter(
            id__in=missing_skill_ids
        )

        total_required_skills = len(matched_skill_ids) + len(missing_skill_ids)

        if total_required_skills > 0:
            skill_match_score = (
                len(matched_skill_ids) / total_required_skills
            ) * 100
        else:
            skill_match_score = 0

        matched_tools = user_tools.filter(
            id__in=target_required_tools.values_list('id', flat=True)
        )

        missing_tools = target_required_tools.exclude(
            id__in=user_tools.values_list('id', flat=True)
        )

        if target_required_tools.count() > 0:
            tool_match_score = (
                matched_tools.count() / target_required_tools.count()
            ) * 100
        else:
            tool_match_score = 0

        feasibility_score = (skill_match_score * 0.7) + (tool_match_score * 0.3)

        if feasibility_score >= 75:
            difficulty_level = "Easy Transition"
            recommendation = (
                f"Transition from {current_role.role_name} to {target_role.role_name} is realistic. "
                "Focus on portfolio projects and interview preparation."
            )
        elif feasibility_score >= 50:
            difficulty_level = "Moderate Transition"
            recommendation = (
                f"Transition from {current_role.role_name} to {target_role.role_name} is possible, "
                "but you need to close the missing skill and tool gaps first."
            )
        else:
            difficulty_level = "Difficult Transition"
            recommendation = (
                f"Transition from {current_role.role_name} to {target_role.role_name} is currently difficult. "
                "Build foundational skills and tools before targeting this role."
            )

        analysis = CareerTransitionAnalysis.objects.create(
            user=request.user,
            current_role=current_role,
            target_role=target_role,
            skill_match_score=round(skill_match_score, 2),
            tool_match_score=round(tool_match_score, 2),
            feasibility_score=round(feasibility_score, 2),
            difficulty_level=difficulty_level,
            missing_skills=", ".join(
                missing_skills.values_list('skill_name', flat=True)
            ),
            missing_tools=", ".join(
                missing_tools.values_list('tool_name', flat=True)
            ),
            recommendation=recommendation
        )

        return redirect(
            'career_transition_result',
            analysis_id=analysis.id
        )

    return render(
        request,
        'career_app/career_transition_analysis.html',
        {'form': form}
    )

@login_required
def career_transition_result(request, analysis_id):
    analysis = CareerTransitionAnalysis.objects.get(
        id=analysis_id,
        user=request.user
    )

    return render(
        request,
        'career_app/career_transition_result.html',
        {'analysis': analysis}
    )
@login_required
def import_dataset(request):
    if not request.user.is_staff:
        return redirect('user_dashboard')

    form = DatasetImportForm(
        request.POST or None,
        request.FILES or None
    )

    if request.method != 'POST' or not form.is_valid():
        return render(
            request,
            'career_app/import_dataset.html',
            {
                'form': form
            }
        )

    dataset_file = request.FILES['dataset_file']

    try:
        workbook = load_workbook(
            dataset_file,
            data_only=True
        )
    except Exception as error:
        messages.error(
            request,
            f"Could not open the Excel file: {error}"
        )

        return render(
            request,
            'career_app/import_dataset.html',
            {
                'form': form
            }
        )

    summary = {
        'created': {
            'roles': 0,
            'skills': 0,
            'tools': 0,
            'role_skills': 0,
            'role_tools': 0,
            'resources': 0,
            'competency_groups': 0,
            'competency_group_members': 0,
            'users': 0,
            'user_profiles': 0,
            'user_profile_skills': 0,
            'user_projects': 0,
            'user_project_skills': 0,
            'user_project_tools': 0,
        },

        'updated': {
            'roles': 0,
            'skills': 0,
            'tools': 0,
            'role_skills': 0,
            'role_tools': 0,
            'resources': 0,
            'competency_groups': 0,
            'competency_group_members': 0,
            'users': 0,
            'user_profiles': 0,
            'user_profile_skills': 0,
            'user_projects': 0,
            'user_project_skills': 0,
            'user_project_tools': 0,
        },

        'duplicates': 0,
        'skipped': 0,
        'warnings': [],
    }

    def clean(value):
        if value is None:
            return ''

        return str(value).strip()

    def add_warning(message):
        # Prevent the result page from becoming enormous.
        if len(summary['warnings']) < 200:
            summary['warnings'].append(message)

    def get_sheet(*possible_names):
        for sheet_name in possible_names:
            if sheet_name in workbook.sheetnames:
                return workbook[sheet_name]

        return None

    def get_headers(sheet):
        header_row = next(
            sheet.iter_rows(
                min_row=1,
                max_row=1,
                values_only=True
            ),
            ()
        )

        return {
            clean(value).lower(): index
            for index, value in enumerate(header_row)
            if clean(value)
        }

    def get_value(row, headers, *possible_headers):
        for header_name in possible_headers:
            index = headers.get(
                header_name.lower()
            )

            if index is not None and index < len(row):
                return clean(row[index])

        return ''

    def to_boolean(value, default=True):
        normalized = clean(value).lower()

        if normalized in [
            '0',
            'false',
            'no',
            'inactive',
        ]:
            return False

        if normalized in [
            '1',
            'true',
            'yes',
            'active',
        ]:
            return True

        return default

    # =====================================================
    # JOB ROLES
    # =====================================================

    sheet = get_sheet(
        'JobRoles',
        'JobRole'
    )

    if sheet:
        headers = get_headers(sheet)

        for index, row in enumerate(
            sheet.iter_rows(
                min_row=2,
                values_only=True
            ),
            start=2
        ):
            role_name = get_value(
                row,
                headers,
                'role_name',
                'job_role',
                'job role',
                'role'
            )

            description = get_value(
                row,
                headers,
                'description',
                'role_description',
                'role description'
            )

            if not role_name:
                summary['skipped'] += 1

                add_warning(
                    f"JobRoles row {index}: "
                    "missing role name."
                )

                continue

            role = JobRole.objects.filter(
                role_name__iexact=role_name
            ).first()

            if not role:
                JobRole.objects.create(
                    role_name=role_name,
                    description=description
                )

                summary['created']['roles'] += 1

            else:
                if (
                    description
                    and role.description != description
                ):
                    role.description = description
                    role.save(
                        update_fields=[
                            'description'
                        ]
                    )

                    summary['updated']['roles'] += 1

                else:
                    summary['duplicates'] += 1

    # =====================================================
    # SKILLS
    # =====================================================

    sheet = get_sheet(
        'Skills',
        'Skill'
    )

    if sheet:
        headers = get_headers(sheet)

        for index, row in enumerate(
            sheet.iter_rows(
                min_row=2,
                values_only=True
            ),
            start=2
        ):
            skill_name = get_value(
                row,
                headers,
                'skill_name',
                'skill name',
                'skill'
            )

            category = get_value(
                row,
                headers,
                'category',
                'skill_category',
                'skill category'
            )

            if not skill_name:
                summary['skipped'] += 1

                add_warning(
                    f"Skills row {index}: "
                    "missing skill name."
                )

                continue

            if skill_name.isdigit():
                summary['skipped'] += 1

                add_warning(
                    f"Skills row {index}: "
                    f"invalid numeric skill '{skill_name}'."
                )

                continue

            skill = Skill.objects.filter(
                skill_name__iexact=skill_name
            ).first()

            if not skill:
                Skill.objects.create(
                    skill_name=skill_name,
                    category=category or 'Other'
                )

                summary['created']['skills'] += 1

            else:
                if (
                    category
                    and skill.category != category
                ):
                    skill.category = category
                    skill.save(
                        update_fields=[
                            'category'
                        ]
                    )

                    summary['updated']['skills'] += 1

                else:
                    summary['duplicates'] += 1

    # =====================================================
    # INDUSTRY TOOLS
    # =====================================================

    sheet = get_sheet(
        'Tools',
        'IndustryTools',
        'IndustryTool'
    )

    if sheet:
        headers = get_headers(sheet)

        for index, row in enumerate(
            sheet.iter_rows(
                min_row=2,
                values_only=True
            ),
            start=2
        ):
            tool_name = get_value(
                row,
                headers,
                'tool_name',
                'tool name',
                'tool'
            )

            category = get_value(
                row,
                headers,
                'category',
                'tool_category',
                'tool category'
            )

            if not tool_name:
                summary['skipped'] += 1

                add_warning(
                    f"Tools row {index}: "
                    "missing tool name."
                )

                continue

            if tool_name.isdigit():
                summary['skipped'] += 1

                add_warning(
                    f"Tools row {index}: "
                    f"invalid numeric tool '{tool_name}'."
                )

                continue

            tool = IndustryTool.objects.filter(
                tool_name__iexact=tool_name
            ).first()

            if not tool:
                IndustryTool.objects.create(
                    tool_name=tool_name,
                    category=category or 'Other'
                )

                summary['created']['tools'] += 1

            else:
                if (
                    category
                    and tool.category != category
                ):
                    tool.category = category
                    tool.save(
                        update_fields=[
                            'category'
                        ]
                    )

                    summary['updated']['tools'] += 1

                else:
                    summary['duplicates'] += 1

    # =====================================================
    # ROLE-SKILL MAPPINGS
    # =====================================================

    sheet = get_sheet(
        'RoleSkills',
        'JobRoleSkills',
        'JobRoleSkill'
    )

    if sheet:
        headers = get_headers(sheet)

        for index, row in enumerate(
            sheet.iter_rows(
                min_row=2,
                values_only=True
            ),
            start=2
        ):
            role_name = get_value(
                row,
                headers,
                'role_name',
                'job_role',
                'job role',
                'role'
            )

            skill_name = get_value(
                row,
                headers,
                'skill_name',
                'skill name',
                'skill'
            )

            importance = get_value(
                row,
                headers,
                'importance',
                'priority'
            ) or 'Medium'

            if not role_name or not skill_name:
                summary['skipped'] += 1

                add_warning(
                    f"RoleSkills row {index}: "
                    "missing role or skill."
                )

                continue

            if importance not in [
                'High',
                'Medium',
                'Low'
            ]:
                summary['skipped'] += 1

                add_warning(
                    f"RoleSkills row {index}: "
                    f"invalid importance '{importance}'."
                )

                continue

            job_role = JobRole.objects.filter(
                role_name__iexact=role_name
            ).first()

            skill = Skill.objects.filter(
                skill_name__iexact=skill_name
            ).first()

            if not job_role:
                summary['skipped'] += 1

                add_warning(
                    f"RoleSkills row {index}: "
                    f"role '{role_name}' not found."
                )

                continue

            if not skill:
                summary['skipped'] += 1

                add_warning(
                    f"RoleSkills row {index}: "
                    f"skill '{skill_name}' not found."
                )

                continue

            mapping = JobRoleSkill.objects.filter(
                job_role=job_role,
                skill=skill
            ).first()

            if not mapping:
                JobRoleSkill.objects.create(
                    job_role=job_role,
                    skill=skill,
                    importance=importance
                )

                summary['created']['role_skills'] += 1

            else:
                if mapping.importance != importance:
                    mapping.importance = importance
                    mapping.save(
                        update_fields=[
                            'importance'
                        ]
                    )

                    summary['updated']['role_skills'] += 1

                else:
                    summary['duplicates'] += 1

    # =====================================================
    # ROLE-TOOL MAPPINGS
    # =====================================================

    sheet = get_sheet(
        'RoleTools',
        'JobRoleTools',
        'JobRoleTool'
    )

    if sheet:
        headers = get_headers(sheet)

        for index, row in enumerate(
            sheet.iter_rows(
                min_row=2,
                values_only=True
            ),
            start=2
        ):
            role_name = get_value(
                row,
                headers,
                'role_name',
                'job_role',
                'job role',
                'role'
            )

            tool_name = get_value(
                row,
                headers,
                'tool_name',
                'tool name',
                'tool'
            )

            importance = get_value(
                row,
                headers,
                'importance',
                'priority'
            ) or 'Medium'

            if not role_name or not tool_name:
                summary['skipped'] += 1

                add_warning(
                    f"RoleTools row {index}: "
                    "missing role or tool."
                )

                continue

            if importance not in [
                'High',
                'Medium',
                'Low'
            ]:
                summary['skipped'] += 1

                add_warning(
                    f"RoleTools row {index}: "
                    f"invalid importance '{importance}'."
                )

                continue

            job_role = JobRole.objects.filter(
                role_name__iexact=role_name
            ).first()

            tool = IndustryTool.objects.filter(
                tool_name__iexact=tool_name
            ).first()

            if not job_role:
                summary['skipped'] += 1

                add_warning(
                    f"RoleTools row {index}: "
                    f"role '{role_name}' not found."
                )

                continue

            if not tool:
                summary['skipped'] += 1

                add_warning(
                    f"RoleTools row {index}: "
                    f"tool '{tool_name}' not found."
                )

                continue

            mapping = JobRoleTool.objects.filter(
                job_role=job_role,
                tool=tool
            ).first()

            if not mapping:
                JobRoleTool.objects.create(
                    job_role=job_role,
                    tool=tool,
                    importance=importance
                )

                summary['created']['role_tools'] += 1

            else:
                if mapping.importance != importance:
                    mapping.importance = importance
                    mapping.save(
                        update_fields=[
                            'importance'
                        ]
                    )

                    summary['updated']['role_tools'] += 1

                else:
                    summary['duplicates'] += 1

    # =====================================================
    # LEARNING RESOURCES
    # =====================================================

    sheet = get_sheet(
        'LearningResources',
        'LearningResource'
    )

    if sheet:
        headers = get_headers(sheet)

        valid_resource_types = [
            'Course',
            'Video',
            'Documentation',
            'Book',
            'Article',
        ]

        for index, row in enumerate(
            sheet.iter_rows(
                min_row=2,
                values_only=True
            ),
            start=2
        ):
            skill_name = get_value(
                row,
                headers,
                'skill_name',
                'skill name',
                'skill'
            )

            title = get_value(
                row,
                headers,
                'title',
                'resource_title',
                'resource title'
            )

            resource_type = get_value(
                row,
                headers,
                'resource_type',
                'resource type',
                'type'
            ) or 'Course'

            url = get_value(
                row,
                headers,
                'url',
                'link',
                'resource_url',
                'resource url'
            )

            if (
                not skill_name
                or not title
                or not url
            ):
                summary['skipped'] += 1

                add_warning(
                    f"LearningResources row {index}: "
                    "missing skill, title or URL."
                )

                continue

            if resource_type not in valid_resource_types:
                summary['skipped'] += 1

                add_warning(
                    f"LearningResources row {index}: "
                    f"invalid resource type "
                    f"'{resource_type}'."
                )

                continue

            skill = Skill.objects.filter(
                skill_name__iexact=skill_name
            ).first()

            if not skill:
                summary['skipped'] += 1

                add_warning(
                    f"LearningResources row {index}: "
                    f"skill '{skill_name}' not found."
                )

                continue

            resource = LearningResource.objects.filter(
                skill=skill,
                title__iexact=title
            ).first()

            if not resource:
                LearningResource.objects.create(
                    skill=skill,
                    title=title,
                    resource_type=resource_type,
                    url=url
                )

                summary['created']['resources'] += 1

            else:
                changed = False

                if resource.resource_type != resource_type:
                    resource.resource_type = resource_type
                    changed = True

                if resource.url != url:
                    resource.url = url
                    changed = True

                if changed:
                    resource.save()

                    summary['updated']['resources'] += 1

                else:
                    summary['duplicates'] += 1

    # =====================================================
    # COMPETENCY GROUPS
    # =====================================================

    sheet = get_sheet(
        'CompetencyGroups',
        'CompetencyGroup'
    )

    if sheet:
        headers = get_headers(sheet)

        for index, row in enumerate(
            sheet.iter_rows(
                min_row=2,
                values_only=True
            ),
            start=2
        ):
            role_name = get_value(
                row,
                headers,
                'role_name',
                'job_role',
                'job role',
                'role'
            )

            group_name = get_value(
                row,
                headers,
                'group_name',
                'group name',
                'competency_group',
                'competency group'
            )

            rule = get_value(
                row,
                headers,
                'rule',
                'group_rule',
                'group rule'
            ).upper()

            if (
                not role_name
                or not group_name
                or not rule
            ):
                summary['skipped'] += 1

                add_warning(
                    f"CompetencyGroups row {index}: "
                    "missing role, group name or rule."
                )

                continue

            if rule not in [
                'ANY_ONE',
                'ALL_REQUIRED'
            ]:
                summary['skipped'] += 1

                add_warning(
                    f"CompetencyGroups row {index}: "
                    f"invalid rule '{rule}'."
                )

                continue

            job_role = JobRole.objects.filter(
                role_name__iexact=role_name
            ).first()

            if not job_role:
                summary['skipped'] += 1

                add_warning(
                    f"CompetencyGroups row {index}: "
                    f"role '{role_name}' not found."
                )

                continue

            group = CompetencyGroup.objects.filter(
                job_role=job_role,
                group_name__iexact=group_name
            ).first()

            if not group:
                CompetencyGroup.objects.create(
                    job_role=job_role,
                    group_name=group_name,
                    rule=rule
                )

                summary['created'][
                    'competency_groups'
                ] += 1

            else:
                if group.rule != rule:
                    group.rule = rule
                    group.save(
                        update_fields=[
                            'rule'
                        ]
                    )

                    summary['updated'][
                        'competency_groups'
                    ] += 1

                else:
                    summary['duplicates'] += 1

    # =====================================================
    # COMPETENCY GROUP MEMBERS
    # =====================================================

    sheet = get_sheet(
        'CompetencyGroupMembers',
        'CompetencyGroupMember'
    )

    if sheet:
        headers = get_headers(sheet)

        for index, row in enumerate(
            sheet.iter_rows(
                min_row=2,
                values_only=True
            ),
            start=2
        ):
            role_name = get_value(
                row,
                headers,
                'role_name',
                'job_role',
                'job role',
                'role'
            )

            group_name = get_value(
                row,
                headers,
                'group_name',
                'group name',
                'competency_group',
                'competency group'
            )

            skill_name = get_value(
                row,
                headers,
                'skill_name',
                'skill name',
                'skill'
            )

            if (
                not role_name
                or not group_name
                or not skill_name
            ):
                summary['skipped'] += 1

                add_warning(
                    f"CompetencyGroupMembers row {index}: "
                    "missing role, group name or skill."
                )

                continue

            job_role = JobRole.objects.filter(
                role_name__iexact=role_name
            ).first()

            if not job_role:
                summary['skipped'] += 1

                add_warning(
                    f"CompetencyGroupMembers row {index}: "
                    f"role '{role_name}' not found."
                )

                continue

            group = CompetencyGroup.objects.filter(
                job_role=job_role,
                group_name__iexact=group_name
            ).first()

            if not group:
                summary['skipped'] += 1

                add_warning(
                    f"CompetencyGroupMembers row {index}: "
                    f"group '{group_name}' not found "
                    f"for role '{role_name}'."
                )

                continue

            job_role_skill = JobRoleSkill.objects.filter(
                job_role=job_role,
                skill__skill_name__iexact=skill_name
            ).select_related(
                'skill'
            ).first()

            if not job_role_skill:
                summary['skipped'] += 1

                add_warning(
                    f"CompetencyGroupMembers row {index}: "
                    f"skill '{skill_name}' is not mapped "
                    f"to role '{role_name}'."
                )

                continue

            member, created = (
                CompetencyGroupMember.objects.get_or_create(
                    group=group,
                    job_role_skill=job_role_skill
                )
            )

            if created:
                summary['created'][
                    'competency_group_members'
                ] += 1

            else:
                summary['duplicates'] += 1

    # =====================================================
    # USERS
    # =====================================================

    sheet = get_sheet(
        'User',
        'Users'
    )

    if sheet:
        headers = get_headers(sheet)

        for index, row in enumerate(
            sheet.iter_rows(
                min_row=2,
                values_only=True
            ),
            start=2
        ):
            username = get_value(
                row,
                headers,
                'username',
                'user_name',
                'user name'
            )

            first_name = get_value(
                row,
                headers,
                'first_name',
                'first name'
            )

            last_name = get_value(
                row,
                headers,
                'last_name',
                'last name'
            )

            email = get_value(
                row,
                headers,
                'email',
                'email_address',
                'email address'
            )

            is_active_value = get_value(
                row,
                headers,
                'is_active',
                'active'
            )

            password = get_value(
                row,
                headers,
                'temporary_password',
                'password'
            ) or 'CareerReady@123'

            username = User.normalize_username(
                username
            )

            if not username or not email:
                summary['skipped'] += 1

                add_warning(
                    f"User row {index}: "
                    "missing username or email."
                )

                continue

            expected_active = to_boolean(
                is_active_value,
                default=True
            )

            user = User.objects.filter(
                username__iexact=username
            ).first()

            if user:
                if (
                    user.is_staff
                    or user.is_superuser
                ):
                    summary['skipped'] += 1

                    add_warning(
                        f"User row {index}: "
                        f"'{username}' is an admin account "
                        "and was not changed."
                    )

                    continue

                email_conflict = User.objects.filter(
                    email__iexact=email
                ).exclude(
                    id=user.id
                ).exists()

                if email_conflict:
                    summary['skipped'] += 1

                    add_warning(
                        f"User row {index}: "
                        f"email '{email}' is already used."
                    )

                    continue

                changed = False

                if user.first_name != first_name:
                    user.first_name = first_name
                    changed = True

                if user.last_name != last_name:
                    user.last_name = last_name
                    changed = True

                if user.email != email:
                    user.email = email
                    changed = True

                if user.is_active != expected_active:
                    user.is_active = expected_active
                    changed = True

                if changed:
                    user.save(
                        update_fields=[
                            'first_name',
                            'last_name',
                            'email',
                            'is_active',
                        ]
                    )

                    summary['updated']['users'] += 1

                else:
                    summary['duplicates'] += 1

                continue

            if User.objects.filter(
                email__iexact=email
            ).exists():
                summary['skipped'] += 1

                add_warning(
                    f"User row {index}: "
                    f"email '{email}' is already used."
                )

                continue

            try:
                User.objects.create_user(
                    username=username,
                    email=email,
                    password=password,
                    first_name=first_name,
                    last_name=last_name,
                    is_active=expected_active,
                    is_staff=False,
                    is_superuser=False
                )

                summary['created']['users'] += 1

            except IntegrityError:
                summary['skipped'] += 1

                add_warning(
                    f"User row {index}: "
                    f"username '{username}' could not "
                    "be created."
                )

    # =====================================================
    # USER PROFILES
    # =====================================================

    profile_source_map = {}

    sheet = get_sheet(
        'UserProfile',
        'UserProfiles'
    )

    if sheet:
        headers = get_headers(sheet)

        for index, row in enumerate(
            sheet.iter_rows(
                min_row=2,
                values_only=True
            ),
            start=2
        ):
            source_profile_id = get_value(
                row,
                headers,
                'id',
                'user_profile_id',
                'profile_id'
            )

            username = get_value(
                row,
                headers,
                'username',
                'user_name',
                'user name'
            )

            full_name = get_value(
                row,
                headers,
                'full_name',
                'full name'
            )

            university = get_value(
                row,
                headers,
                'university'
            )

            degree = get_value(
                row,
                headers,
                'degree'
            )

            year_of_study = get_value(
                row,
                headers,
                'year_of_study',
                'year of study'
            )

            github_link = get_value(
                row,
                headers,
                'github_link',
                'github_url',
                'github'
            )

            linkedin_link = get_value(
                row,
                headers,
                'linkedin_link',
                'linkedin_url',
                'linkedin'
            )

            resume_valid_value = get_value(
                row,
                headers,
                'is_resume_valid',
                'resume_valid'
            )

            if not username:
                summary['skipped'] += 1

                add_warning(
                    f"UserProfile row {index}: "
                    "missing username."
                )

                continue

            user = User.objects.filter(
                username__iexact=username
            ).first()

            if not user:
                summary['skipped'] += 1

                add_warning(
                    f"UserProfile row {index}: "
                    f"user '{username}' not found."
                )

                continue

            default_full_name = (
                user.get_full_name().strip()
                or user.username
            )

            values = {
                'full_name': (
                    full_name
                    or default_full_name
                ),
                'university': university or None,
                'degree': degree or None,
                'year_of_study': (
                    year_of_study or None
                ),
                'github_link': github_link or None,
                'linkedin_link': (
                    linkedin_link or None
                ),
                'is_resume_valid': to_boolean(
                    resume_valid_value,
                    default=False
                ),
            }

            profile = UserProfile.objects.filter(
                user=user
            ).first()

            if not profile:
                profile = UserProfile.objects.create(
                    user=user,
                    **values
                )

                summary['created'][
                    'user_profiles'
                ] += 1

            else:
                changed = False

                for field_name, value in values.items():
                    if getattr(
                        profile,
                        field_name
                    ) != value:
                        setattr(
                            profile,
                            field_name,
                            value
                        )

                        changed = True

                if changed:
                    profile.save()

                    summary['updated'][
                        'user_profiles'
                    ] += 1

                else:
                    summary['duplicates'] += 1

            if source_profile_id:
                profile_source_map[
                    source_profile_id
                ] = profile

    # =====================================================
    # USER PROFILE SKILLS
    # =====================================================

    sheet = get_sheet(
        'UserProfile_Skills',
        'UserProfileSkills'
    )

    if sheet:
        headers = get_headers(sheet)

        for index, row in enumerate(
            sheet.iter_rows(
                min_row=2,
                values_only=True
            ),
            start=2
        ):
            source_profile_id = get_value(
                row,
                headers,
                'user_profile_id',
                'profile_id'
            )

            username = get_value(
                row,
                headers,
                'username',
                'user_name'
            )

            skill_name = get_value(
                row,
                headers,
                'skill_name',
                'skill name',
                'skill'
            )

            profile = None

            if source_profile_id:
                profile = profile_source_map.get(
                    source_profile_id
                )

            if not profile and username:
                profile = UserProfile.objects.filter(
                    user__username__iexact=username
                ).first()

            skill = Skill.objects.filter(
                skill_name__iexact=skill_name
            ).first()

            if not profile or not skill:
                summary['skipped'] += 1

                add_warning(
                    f"UserProfile_Skills row {index}: "
                    "profile or skill not found."
                )

                continue

            if profile.manual_skills.filter(
                id=skill.id
            ).exists():
                summary['duplicates'] += 1

            else:
                profile.manual_skills.add(
                    skill
                )

                summary['created'][
                    'user_profile_skills'
                ] += 1

    # =====================================================
    # USER PROJECTS
    # =====================================================

    project_source_map = {}

    sheet = get_sheet(
        'UserProject',
        'UserProjects'
    )

    if sheet:
        headers = get_headers(sheet)

        for index, row in enumerate(
            sheet.iter_rows(
                min_row=2,
                values_only=True
            ),
            start=2
        ):
            source_project_id = get_value(
                row,
                headers,
                'id',
                'user_project_id',
                'project_id'
            )

            username = get_value(
                row,
                headers,
                'username',
                'user_name'
            )

            title = get_value(
                row,
                headers,
                'title',
                'project_title',
                'project title'
            )

            description = get_value(
                row,
                headers,
                'description',
                'project_description',
                'project description'
            )

            project_url = get_value(
                row,
                headers,
                'project_url',
                'project link',
                'url'
            )

            github_url = get_value(
                row,
                headers,
                'github_url',
                'github_link',
                'github'
            )

            if not username or not title:
                summary['skipped'] += 1

                add_warning(
                    f"UserProject row {index}: "
                    "missing username or project title."
                )

                continue

            user = User.objects.filter(
                username__iexact=username
            ).first()

            if not user:
                summary['skipped'] += 1

                add_warning(
                    f"UserProject row {index}: "
                    f"user '{username}' not found."
                )

                continue

            project = UserProject.objects.filter(
                user=user,
                title__iexact=title
            ).first()

            if not project:
                project = UserProject.objects.create(
                    user=user,
                    title=title,
                    description=description,
                    project_url=project_url or None,
                    github_url=github_url or None
                )

                summary['created'][
                    'user_projects'
                ] += 1

            else:
                changed = False

                expected_values = {
                    'description': description,
                    'project_url': (
                        project_url or None
                    ),
                    'github_url': (
                        github_url or None
                    ),
                }

                for field_name, value in expected_values.items():
                    if getattr(
                        project,
                        field_name
                    ) != value:
                        setattr(
                            project,
                            field_name,
                            value
                        )

                        changed = True

                if changed:
                    project.save()

                    summary['updated'][
                        'user_projects'
                    ] += 1

                else:
                    summary['duplicates'] += 1

            if source_project_id:
                project_source_map[
                    source_project_id
                ] = project

    # =====================================================
    # USER PROJECT SKILLS
    # =====================================================

    sheet = get_sheet(
        'UserProject_Skills',
        'UserProjectSkills'
    )

    if sheet:
        headers = get_headers(sheet)

        for index, row in enumerate(
            sheet.iter_rows(
                min_row=2,
                values_only=True
            ),
            start=2
        ):
            source_project_id = get_value(
                row,
                headers,
                'user_project_id',
                'project_id'
            )

            skill_name = get_value(
                row,
                headers,
                'skill_name',
                'skill name',
                'skill'
            )

            project = project_source_map.get(
                source_project_id
            )

            skill = Skill.objects.filter(
                skill_name__iexact=skill_name
            ).first()

            if not project or not skill:
                summary['skipped'] += 1

                add_warning(
                    f"UserProject_Skills row {index}: "
                    "project or skill not found."
                )

                continue

            if project.skills_used.filter(
                id=skill.id
            ).exists():
                summary['duplicates'] += 1

            else:
                project.skills_used.add(
                    skill
                )

                summary['created'][
                    'user_project_skills'
                ] += 1

    # =====================================================
    # USER PROJECT TOOLS
    # =====================================================

    sheet = get_sheet(
        'UserProject_Tools',
        'UserProjectTools'
    )

    if sheet:
        headers = get_headers(sheet)

        for index, row in enumerate(
            sheet.iter_rows(
                min_row=2,
                values_only=True
            ),
            start=2
        ):
            source_project_id = get_value(
                row,
                headers,
                'user_project_id',
                'project_id'
            )

            tool_name = get_value(
                row,
                headers,
                'tool_name',
                'tool name',
                'tool'
            )

            project = project_source_map.get(
                source_project_id
            )

            tool = IndustryTool.objects.filter(
                tool_name__iexact=tool_name
            ).first()

            if not project or not tool:
                summary['skipped'] += 1

                add_warning(
                    f"UserProject_Tools row {index}: "
                    "project or tool not found."
                )

                continue

            if project.tools_used.filter(
                id=tool.id
            ).exists():
                summary['duplicates'] += 1

            else:
                project.tools_used.add(
                    tool
                )

                summary['created'][
                    'user_project_tools'
                ] += 1

    created_total = sum(
        summary['created'].values()
    )

    updated_total = sum(
        summary['updated'].values()
    )

    messages.success(
        request,
        "Dataset import completed. "
        f"Created: {created_total}, "
        f"Updated: {updated_total}, "
        f"Duplicates: {summary['duplicates']}, "
        f"Skipped: {summary['skipped']}."
    )

    return render(
        request,
        'career_app/import_dataset.html',
        {
            'form': DatasetImportForm(),
            'summary': summary
        }
    )
@login_required
def add_competency_group(request):
    if not request.user.is_staff:
        return redirect('user_dashboard')

    form = CompetencyGroupForm(request.POST or None)

    if request.method == 'POST' and form.is_valid():
        form.save()
        return redirect('view_competency_groups')

    return render(
        request,
        'career_app/add_competency_group.html',
        {'form': form}
    )


@login_required
def view_competency_groups(request):
    if not request.user.is_staff:
        return redirect('user_dashboard')

    member_queryset = CompetencyGroupMember.objects.select_related(
        'job_role_skill',
        'job_role_skill__skill',
        'job_role_skill__job_role'
    ).order_by(
        'job_role_skill__skill__skill_name'
    )

    groups_queryset = CompetencyGroup.objects.select_related(
        'job_role'
    ).prefetch_related(
        Prefetch(
            'members',
            queryset=member_queryset
        )
    ).order_by(
        'job_role__role_name',
        'group_name'
    )

    paginator = Paginator(
        groups_queryset,
        25
    )

    page_number = request.GET.get('page')

    groups = paginator.get_page(
        page_number
    )

    return render(
        request,
        'career_app/view_competency_groups.html',
        {
            'groups': groups
        }
    )
@login_required
def add_competency_group_members(request):
    if not request.user.is_staff:
        return redirect('user_dashboard')

    selected_group = None
    group_id = request.GET.get('group') or request.POST.get('group')

    if group_id:
        selected_group = CompetencyGroup.objects.filter(id=group_id).first()

    if request.method == 'POST':
        form = CompetencyGroupMemberForm(
            request.POST,
            selected_group=selected_group
        )

        if form.is_valid():
            group = form.cleaned_data['group']
            job_role_skills = form.cleaned_data['job_role_skills']

            CompetencyGroupMember.objects.filter(group=group).delete()

            for job_role_skill in job_role_skills:
                CompetencyGroupMember.objects.get_or_create(
                    group=group,
                    job_role_skill=job_role_skill
                )

            return redirect('view_competency_groups')
    else:
        form = CompetencyGroupMemberForm(
            selected_group=selected_group
        )

    return render(
        request,
        'career_app/add_competency_group_members.html',
        {
            'form': form,
            'selected_group': selected_group
        }
    )

@login_required
def edit_competency_group(request, group_id):
    if not request.user.is_staff:
        return redirect('user_dashboard')

    group = CompetencyGroup.objects.get(id=group_id)

    form = CompetencyGroupForm(
        request.POST or None,
        instance=group
    )

    if request.method == 'POST' and form.is_valid():
        form.save()
        return redirect('view_competency_groups')

    return render(
        request,
        'career_app/edit_competency_group.html',
        {
            'form': form,
            'group': group
        }
    )


@login_required
def delete_competency_group(request, group_id):
    if not request.user.is_staff:
        return redirect('user_dashboard')

    group = CompetencyGroup.objects.get(id=group_id)

    if request.method == 'POST':
        group.delete()
        return redirect('view_competency_groups')

    return render(
        request,
        'career_app/delete_competency_group.html',
        {'group': group}
    )

@login_required
def platform_analytics(request):
    if not request.user.is_superuser:
        return redirect('user_dashboard')

    context = {
        'total_users': User.objects.filter(is_staff=False, is_superuser=False).count(),
        'total_admins': User.objects.filter(is_staff=True, is_superuser=False).count(),
        'total_job_roles': JobRole.objects.count(),
        'total_skills': Skill.objects.count(),
        'total_tools': IndustryTool.objects.count(),
        'total_learning_resources': LearningResource.objects.count(),
        'total_projects': UserProject.objects.count(),
        'total_readiness_assessments': ReadinessAssessment.objects.count(),
        'total_bottleneck_reports': EmployabilityBottleneck.objects.count(),
        'total_transition_analyses': CareerTransitionAnalysis.objects.count(),
        'pending_admin_requests': AdminRequest.objects.filter(status='Pending').count(),
        'approved_admin_requests': AdminRequest.objects.filter(status='Approved').count(),
        'rejected_admin_requests': AdminRequest.objects.filter(status='Rejected').count(),
        'used_invites': AdminInviteCode.objects.filter(is_used=True).count(),
        'unused_invites': AdminInviteCode.objects.filter(is_used=False).count(),
    }

    return render(
        request,
        'career_app/platform_analytics.html',
        context
    )


def generate_interview_questions(session):
    """
    Generate personalised interview questions using:

    - Selected project
    - Target job role
    - Project skills and tools
    - User profile skills and tools
    - Competency groups
    - Readiness assessment
    - Employability bottleneck

    Important competency rule:

    - ANY_ONE groups produce only one interview question.
    - If the user knows one option, the whole group is satisfied.
    - The remaining options are not treated as separate weaknesses.
    - ALL_REQUIRED groups can generate questions for individual missing skills.
    """

    # Do not create duplicate questions.
    if session.questions.exists():
        return session.questions.order_by(
            'display_order',
            'id'
        )

    user = session.user
    job_role = session.job_role
    project = session.project

    # -------------------------------------------------
    # Latest assessment and bottleneck
    # -------------------------------------------------

    latest_assessment = ReadinessAssessment.objects.filter(
        user=user,
        job_role=job_role
    ).order_by(
        '-created_at'
    ).first()

    latest_bottleneck = EmployabilityBottleneck.objects.filter(
        user=user,
        job_role=job_role
    ).order_by(
        '-created_at'
    ).first()

    # -------------------------------------------------
    # Role skills and tools
    # -------------------------------------------------

    role_skills = list(
        JobRoleSkill.objects.filter(
            job_role=job_role
        ).select_related(
            'skill'
        ).order_by(
            'importance',
            'skill__skill_name'
        )
    )

    role_tools = list(
        JobRoleTool.objects.filter(
            job_role=job_role
        ).select_related(
            'tool'
        ).order_by(
            'importance',
            'tool__tool_name'
        )
    )

    # -------------------------------------------------
    # Project evidence
    # -------------------------------------------------

    project_skill_ids = set(
        project.skills_used.values_list(
            'id',
            flat=True
        )
    )

    project_tool_ids = set(
        project.tools_used.values_list(
            'id',
            flat=True
        )
    )

    # Begin with project evidence.
    user_skill_ids = set(project_skill_ids)
    user_tool_ids = set(project_tool_ids)

    # -------------------------------------------------
    # User profile evidence
    # -------------------------------------------------

    try:
        user_profile = user.userprofile
    except UserProfile.DoesNotExist:
        user_profile = None

    if user_profile:
        user_skill_ids.update(
            user_profile.manual_skills.values_list(
                'id',
                flat=True
            )
        )

        user_skill_ids.update(
            user_profile.extracted_skills.values_list(
                'id',
                flat=True
            )
        )

        user_tool_ids.update(
            user_profile.manual_tools.values_list(
                'id',
                flat=True
            )
        )

    professional_skill_names = {
        'communication',
        'teamwork',
        'collaboration',
        'leadership',
        'adaptability',
        'time management',
        'problem solving',
        'critical thinking',
        'creativity',
        'presentation skills',
        'decision making',
        'conflict resolution',
        'attention to detail',
    }

    question_data = []

    # -------------------------------------------------
    # Helper for adding unique questions
    # -------------------------------------------------

    def add_question(
        question_text,
        question_type,
        difficulty='MEDIUM',
        expected_skill=None,
        expected_tool=None,
        competency_group=None
    ):
        normalized_text = ' '.join(
            question_text.lower().split()
        )

        existing_texts = {
            ' '.join(
                item['question_text'].lower().split()
            )
            for item in question_data
        }

        if normalized_text in existing_texts:
            return

        question_data.append({
            'question_text': question_text,
            'question_type': question_type,
            'difficulty': difficulty,
            'expected_skill': expected_skill,
            'expected_tool': expected_tool,
            'competency_group': competency_group,
        })

    # =================================================
    # 1. Project introduction
    # =================================================

    add_question(
        question_text=(
            f"Please introduce your project '{project.title}'. "
            f"What problem does it solve, who are its intended users, "
            f"and what was your personal contribution?"
        ),
        question_type='PROJECT',
        difficulty='EASY'
    )

    # =================================================
    # 2. Project architecture
    # =================================================

    add_question(
        question_text=(
            f"Describe the technical architecture of '{project.title}'. "
            f"Explain the main components, how data moves through the "
            f"system, and why you selected this design."
        ),
        question_type='SYSTEM_DESIGN',
        difficulty='MEDIUM'
    )

    # =================================================
    # 3. Project challenge
    # =================================================

    add_question(
        question_text=(
            f"What was the most difficult technical challenge you faced "
            f"while developing '{project.title}'? Explain how you diagnosed "
            f"the problem, the solution you implemented, and what you learned."
        ),
        question_type='PROJECT',
        difficulty='MEDIUM'
    )

    # =================================================
    # 4. Matched technical project skills
    # =================================================

    matched_technical_skills = [
        role_skill
        for role_skill in role_skills
        if (
            role_skill.skill_id in project_skill_ids
            and role_skill.skill.skill_name.strip().lower()
            not in professional_skill_names
        )
    ]

    for role_skill in matched_technical_skills[:2]:
        skill = role_skill.skill

        difficulty = (
            'HARD'
            if role_skill.importance == 'High'
            else 'MEDIUM'
        )

        add_question(
            question_text=(
                f"You listed {skill.skill_name} as a skill used in "
                f"'{project.title}'. Describe a specific feature where you "
                f"applied it. What technical decisions did you make, and "
                f"how did you test that the implementation worked correctly?"
            ),
            question_type='TECHNICAL',
            difficulty=difficulty,
            expected_skill=skill
        )

    # =================================================
    # 5. Matched professional skill
    # =================================================

    matched_professional_skills = [
        role_skill
        for role_skill in role_skills
        if (
            role_skill.skill_id in project_skill_ids
            and role_skill.skill.skill_name.strip().lower()
            in professional_skill_names
        )
    ]

    for role_skill in matched_professional_skills[:1]:
        skill = role_skill.skill

        add_question(
            question_text=(
                f"Give an example of how you demonstrated "
                f"{skill.skill_name} while working on "
                f"'{project.title}'. Describe the situation, the action "
                f"you took, and the final outcome."
            ),
            question_type='BEHAVIOURAL',
            difficulty='MEDIUM',
            expected_skill=skill
        )

    # =================================================
    # 6. Matched project tools
    # =================================================

    matched_project_tools = [
        role_tool
        for role_tool in role_tools
        if role_tool.tool_id in project_tool_ids
    ]

    for role_tool in matched_project_tools[:1]:
        tool = role_tool.tool

        difficulty = (
            'HARD'
            if role_tool.importance == 'High'
            else 'MEDIUM'
        )

        add_question(
            question_text=(
                f"How did you use {tool.tool_name} in "
                f"'{project.title}'? Explain why it was selected, "
                f"how it supported development, and one limitation or "
                f"problem you encountered while using it."
            ),
            question_type='TOOL',
            difficulty=difficulty,
            expected_tool=tool
        )

    # =================================================
    # 7. Load competency groups
    # =================================================

    competency_groups = list(
        CompetencyGroup.objects.filter(
            job_role=job_role
        ).prefetch_related(
            'members__job_role_skill__skill'
        ).order_by(
            'group_name'
        )
    )

    grouped_skill_ids = set()

    satisfied_any_one_groups = []
    unsatisfied_any_one_groups = []
    missing_all_required_skills = []

    for group in competency_groups:
        group_members = [
            member
            for member in group.members.all()
            if member.job_role_skill
        ]

        if not group_members:
            continue

        group_skill_ids = {
            member.job_role_skill.skill_id
            for member in group_members
        }

        grouped_skill_ids.update(
            group_skill_ids
        )

        matched_skill_ids = group_skill_ids.intersection(
            user_skill_ids
        )

        # ---------------------------------------------
        # ANY_ONE
        # ---------------------------------------------

        if group.rule == 'ANY_ONE':
            if matched_skill_ids:
                satisfied_any_one_groups.append({
                    'group': group,
                    'members': group_members,
                    'matched_skill_ids': matched_skill_ids,
                })
            else:
                unsatisfied_any_one_groups.append({
                    'group': group,
                    'members': group_members,
                })

        # ---------------------------------------------
        # ALL_REQUIRED
        # ---------------------------------------------

        elif group.rule == 'ALL_REQUIRED':
            for member in group_members:
                role_skill = member.job_role_skill

                if role_skill.skill_id not in user_skill_ids:
                    missing_all_required_skills.append(
                        role_skill
                    )

    # =================================================
    # 8. Ask one question for a satisfied ANY_ONE group
    # =================================================

    for group_data in satisfied_any_one_groups[:1]:
        group = group_data['group']
        group_members = group_data['members']
        matched_skill_ids = group_data['matched_skill_ids']

        matched_skills = [
            member.job_role_skill.skill
            for member in group_members
            if (
                member.job_role_skill.skill_id
                in matched_skill_ids
            )
        ]

        matched_skill_names = ', '.join(
            skill.skill_name
            for skill in matched_skills
        )

        representative_skill = (
            matched_skills[0]
            if matched_skills
            else None
        )

        add_question(
            question_text=(
                f"You satisfy the competency group "
                f"'{group.group_name}' through {matched_skill_names}. "
                f"Which of these competencies are you strongest in? "
                f"Explain how you applied it in '{project.title}' "
                f"and why it was suitable."
            ),
            question_type='COMPETENCY',
            difficulty='MEDIUM',
            expected_skill=representative_skill,
            competency_group=group
        )

    # =================================================
    # 9. Ask one question per missing ANY_ONE group
    # =================================================

    for group_data in unsatisfied_any_one_groups[:2]:
        group = group_data['group']
        group_members = group_data['members']

        skills = [
            member.job_role_skill.skill
            for member in group_members
        ]

        if not skills:
            continue

        skill_names = ', '.join(
            skill.skill_name
            for skill in skills
        )

        representative_skill = skills[0]

        add_question(
            question_text=(
                f"The competency group '{group.group_name}' can be "
                f"satisfied by learning any one of these options: "
                f"{skill_names}. Which option would you choose for "
                f"'{project.title}', and how would you apply it in a "
                f"production environment?"
            ),
            question_type='WEAKNESS',
            difficulty='HARD',
            expected_skill=representative_skill,
            competency_group=group
        )

    # Important:
    # AWS, Azure and Google Cloud Platform will now generate only one
    # question if they belong to the same ANY_ONE group.

    # =================================================
    # 10. Missing ALL_REQUIRED skills
    # =================================================

    technical_all_required_skills = [
        role_skill
        for role_skill in missing_all_required_skills
        if (
            role_skill.skill.skill_name.strip().lower()
            not in professional_skill_names
        )
    ]

    high_priority_all_required = [
        role_skill
        for role_skill in technical_all_required_skills
        if role_skill.importance == 'High'
    ]

    all_required_candidates = (
        high_priority_all_required
        if high_priority_all_required
        else technical_all_required_skills
    )

    for role_skill in all_required_candidates[:2]:
        skill = role_skill.skill

        add_question(
            question_text=(
                f"{skill.skill_name} is an important required competency "
                f"for the {job_role.role_name} role. Explain the underlying "
                f"concept and describe how you would apply it if "
                f"'{project.title}' needed to be extended for production use."
            ),
            question_type='WEAKNESS',
            difficulty='HARD',
            expected_skill=skill
        )

    # =================================================
    # 11. Ungrouped missing skills
    # =================================================

    ungrouped_missing_skills = [
        role_skill
        for role_skill in role_skills
        if (
            role_skill.skill_id not in grouped_skill_ids
            and role_skill.skill_id not in user_skill_ids
            and role_skill.skill.skill_name.strip().lower()
            not in professional_skill_names
        )
    ]

    high_priority_ungrouped = [
        role_skill
        for role_skill in ungrouped_missing_skills
        if role_skill.importance == 'High'
    ]

    ungrouped_candidates = (
        high_priority_ungrouped
        if high_priority_ungrouped
        else ungrouped_missing_skills
    )

    for role_skill in ungrouped_candidates[:1]:
        skill = role_skill.skill

        add_question(
            question_text=(
                f"{skill.skill_name} is relevant to the "
                f"{job_role.role_name} role. Explain the underlying concept "
                f"and describe how you would apply it to "
                f"'{project.title}' in production."
            ),
            question_type='WEAKNESS',
            difficulty='HARD',
            expected_skill=skill
        )

    # =================================================
    # 12. Missing role tool
    # =================================================

    missing_role_tools = [
        role_tool
        for role_tool in role_tools
        if role_tool.tool_id not in user_tool_ids
    ]

    high_priority_missing_tools = [
        role_tool
        for role_tool in missing_role_tools
        if role_tool.importance == 'High'
    ]

    tool_candidates = (
        high_priority_missing_tools
        if high_priority_missing_tools
        else missing_role_tools
    )

    for role_tool in tool_candidates[:1]:
        tool = role_tool.tool

        add_question(
            question_text=(
                f"The {job_role.role_name} role commonly requires "
                f"{tool.tool_name}. How could this tool be introduced into "
                f"'{project.title}'? Explain its practical benefit and the "
                f"steps you would take to integrate it."
            ),
            question_type='TOOL',
            difficulty='MEDIUM',
            expected_tool=tool
        )

    # =================================================
    # 13. Readiness assessment question
    # =================================================

    if latest_assessment:
        add_question(
            question_text=(
                f"Your latest readiness assessment for "
                f"{job_role.role_name} produced an academic readiness score "
                f"of {latest_assessment.academic_score}% and an industry "
                f"readiness score of {latest_assessment.industry_score}%. "
                f"Which area needs the most improvement, and what practical "
                f"steps would you take to improve it?"
            ),
            question_type='WEAKNESS',
            difficulty='MEDIUM'
        )

    # =================================================
    # 14. Employability bottleneck question
    # =================================================

    if latest_bottleneck:
        bottleneck_name = getattr(
            latest_bottleneck,
            'main_bottleneck',
            'Employability readiness'
        )

        add_question(
            question_text=(
                f"Your employability analysis identified "
                f"'{bottleneck_name}' as an important development area. "
                f"How would you address this bottleneck, and what evidence "
                f"could you produce to demonstrate improvement?"
            ),
            question_type='WEAKNESS',
            difficulty='MEDIUM'
        )

    # =================================================
    # 15. Behavioural fallback
    # =================================================

    add_question(
        question_text=(
            f"Tell me about a time you received critical feedback while "
            f"working on '{project.title}' or another software project. "
            f"How did you respond, what action did you take, and what was "
            f"the final result?"
        ),
        question_type='BEHAVIOURAL',
        difficulty='MEDIUM'
    )

    # =================================================
    # 16. Production readiness
    # =================================================

    add_question(
        question_text=(
            f"Imagine '{project.title}' is going to be used by thousands "
            f"of users. What changes would you make to improve its security, "
            f"performance, scalability, testing and maintainability?"
        ),
        question_type='SYSTEM_DESIGN',
        difficulty='HARD'
    )

    # =================================================
    # Save maximum 10 questions
    # =================================================

    selected_questions = question_data[:10]

    for index, item in enumerate(
        selected_questions,
        start=1
    ):
        InterviewQuestion.objects.create(
            session=session,
            question_type=item['question_type'],
            question_text=item['question_text'],
            difficulty=item['difficulty'],
            display_order=index,
            expected_skill=item['expected_skill'],
            expected_tool=item['expected_tool'],
            competency_group=item['competency_group']
        )

    return session.questions.order_by(
        'display_order',
        'id'
    )
@login_required
def interview_setup(request):
    user_projects = UserProject.objects.filter(user=request.user)

    if not user_projects.exists():
        messages.warning(
            request,
            'You must add at least one project before starting a project-based interview.'
        )
        return redirect('add_project')

    if request.method == 'POST':
        form = InterviewSetupForm(
            request.POST,
            user=request.user
        )

        if form.is_valid():
            interview_session = form.save(commit=False)
            selected_project = form.cleaned_data['project']

            if selected_project.user != request.user:
                messages.error(
                    request,
                    'You cannot use another user’s project.'
                )
                return redirect('interview_setup')

            interview_session.user = request.user
            interview_session.status = 'CREATED'
            interview_session.save()

            # Generate and save interview questions
            generate_interview_questions(interview_session)

            if not interview_session.questions.exists():
                messages.error(
                    request,
                    'The interview session was created, but questions could not be generated.'
                )
                interview_session.delete()
                return redirect('interview_setup')

            interview_session.status = 'IN_PROGRESS'
            interview_session.save(update_fields=['status'])

            messages.success(
                request,
                'Interview session and personalised questions created successfully.'
            )

            return redirect(
                'interview_session',
                session_id=interview_session.id
            )

    else:
        form = InterviewSetupForm(user=request.user)

    context = {
        'form': form,
        'project_count': user_projects.count(),
    }

    return render(
        request,
        'career_app/interview_setup.html',
        context
    )
def evaluate_interview_session(
    session,
    method='hybrid'
):
    """
    Evaluate all answers belonging to one interview session.

    Supported methods:

    rule:
        Use only the rule-based evaluator.

    ai:
        Use only the AI evaluator.
        If AI evaluation fails, raise the error.

    hybrid:
        Try AI evaluation first.
        If any AI evaluation fails, re-evaluate the complete
        session using the rule-based evaluator.

    The method actually used is stored in
    InterviewSession.evaluation_method.
    """

    allowed_methods = {
        'rule',
        'ai',
        'hybrid',
    }

    method = str(
        method or 'hybrid'
    ).strip().lower()

    if method not in allowed_methods:
        raise ValueError(
            "Invalid evaluation method. "
            "Use 'rule', 'ai' or 'hybrid'."
        )

    answers = list(
        InterviewAnswer.objects.filter(
            question__session=session
        ).select_related(
            'question',
            'question__expected_skill',
            'question__expected_tool',
            'question__competency_group',
            'question__session',
            'question__session__project',
            'question__session__job_role',
        ).prefetch_related(
            'question__session__project__skills_used',
            'question__session__project__tools_used',
        ).order_by(
            'question__display_order',
            'question__id',
        )
    )

    if not answers:
        raise ValueError(
            'No interview answers were found for this session.'
        )

    evaluation_method_used = None

    # ---------------------------------------------
    # Rule-based evaluation only
    # ---------------------------------------------

    if method == 'rule':
        for answer in answers:
            evaluate_answer(answer)

        evaluation_method_used = 'RULE_BASED'

    # ---------------------------------------------
    # AI evaluation only
    # ---------------------------------------------

    elif method == 'ai':
        for answer in answers:
            evaluate_answer_with_ai(answer)

        evaluation_method_used = 'AI_POWERED'

    # ---------------------------------------------
    # Hybrid evaluation
    # ---------------------------------------------

    else:
        try:
            for answer in answers:
                evaluate_answer_with_ai(answer)

            evaluation_method_used = 'AI_POWERED'

        except AIInterviewEvaluationError as error:
            logger.warning(
                'AI evaluation failed for interview session %s. '
                'The complete session will be evaluated using '
                'the rule-based fallback. Error: %s',
                session.id,
                error,
            )

            # Re-evaluate every answer using the rule-based evaluator.
            # This avoids mixing AI and rule-based scores in one session.
            for answer in answers:
                evaluate_answer(answer)

            evaluation_method_used = 'RULE_BASED_FALLBACK'

    # ---------------------------------------------
    # Calculate average session score
    # ---------------------------------------------

    average_score = InterviewAnswer.objects.filter(
        question__session=session,
        overall_score__isnull=False,
    ).aggregate(
        average_score=Avg('overall_score')
    )['average_score']

    session.overall_score = (
        round(average_score, 2)
        if average_score is not None
        else 0.0
    )

    session.status = 'COMPLETED'
    session.completed_at = timezone.now()
    session.evaluation_method = evaluation_method_used
    

    session.save(
        update_fields=[
            'overall_score',
            'status',
            'completed_at',
            'evaluation_method',
        ]
    )

    logger.info(
        'Interview session %s evaluated using %s. '
        'Overall score: %s',
        session.id,
        evaluation_method_used,
        session.overall_score,
    )

    return session

@login_required
def interview_session(request, session_id):
    session = get_object_or_404(
        InterviewSession.objects.select_related(
            "job_role",
            "project",
        ),
        id=session_id,
        user=request.user,
    )

    questions = list(
        InterviewQuestion.objects.filter(
            session=session
        ).select_related(
            "expected_skill",
            "expected_tool",
            "competency_group",
        ).order_by(
            "display_order",
            "id",
        )
    )

    total_questions = len(questions)

    if total_questions == 0:
        messages.error(
            request,
            "No interview questions were found for this session.",
        )
        return redirect("interview_setup")

    # If already evaluated, do not allow answer editing.
    if (
        session.status == "COMPLETED"
        and session.evaluation_method
    ):
        return redirect(
            "interview_results",
            session_id=session.id,
        )

    try:
        requested_position = int(
            request.GET.get("question", 1)
        )
    except (TypeError, ValueError):
        requested_position = 1

    requested_position = max(
        1,
        min(requested_position, total_questions),
    )

    current_position = requested_position
    current_question = questions[current_position - 1]

    existing_answer = InterviewAnswer.objects.filter(
        question=current_question
    ).first()

    if request.method == "POST":
        question_id = request.POST.get("question_id")

        posted_question = get_object_or_404(
            InterviewQuestion,
            id=question_id,
            session=session,
        )

        existing_answer = InterviewAnswer.objects.filter(
            question=posted_question
        ).first()

        form = InterviewAnswerForm(
            request.POST,
            instance=existing_answer,
        )

        if form.is_valid():
            interview_answer = form.save(
                commit=False
            )
            interview_answer.question = posted_question
            interview_answer.save()

            posted_position = next(
                (
                    index
                    for index, question in enumerate(
                        questions,
                        start=1,
                    )
                    if question.id == posted_question.id
                ),
                current_position,
            )

            is_last_question = (
                posted_position == total_questions
            )

            # -----------------------------------------
            # Final question
            # -----------------------------------------

            if is_last_question:
                answered_question_ids = set(
                    InterviewAnswer.objects.filter(
                        question__session=session
                    ).exclude(
                        answer_text__isnull=True
                    ).exclude(
                        answer_text__exact=""
                    ).values_list(
                        "question_id",
                        flat=True,
                    )
                )

                first_unanswered_position = None

                for index, question in enumerate(
                    questions,
                    start=1,
                ):
                    if question.id not in answered_question_ids:
                        first_unanswered_position = index
                        break

                    saved_answer = InterviewAnswer.objects.filter(
                        question=question
                    ).first()

                    if (
                        saved_answer is None
                        or not saved_answer.answer_text.strip()
                    ):
                        first_unanswered_position = index
                        break

                if first_unanswered_position is not None:
                    messages.warning(
                        request,
                        "Please answer all interview questions "
                        "before finishing.",
                    )

                    interview_url = reverse(
                        "interview_session",
                        args=[session.id],
                    )

                    return redirect(
                        f"{interview_url}"
                        f"?question={first_unanswered_position}"
                    )

                # Answers are complete, but evaluation has not started.
                # The user will choose Rule-Based or AI evaluation next.
                session.status = "COMPLETED"
                session.completed_at = timezone.now()
                session.evaluation_method = None
                session.overall_score = None

                session.save(
                    update_fields=[
                        "status",
                        "completed_at",
                        "evaluation_method",
                        "overall_score",
                    ]
                )

                return redirect(
                    "choose_interview_evaluation",
                    session_id=session.id,
                )

            # -----------------------------------------
            # Not the final question
            # -----------------------------------------

            next_position = posted_position + 1

            messages.success(
                request,
                "Answer saved successfully.",
            )

            interview_url = reverse(
                "interview_session",
                args=[session.id],
            )

            return redirect(
                f"{interview_url}"
                f"?question={next_position}"
            )

    else:
        form = InterviewAnswerForm(
            instance=existing_answer
        )

    answered_count = InterviewAnswer.objects.filter(
        question__session=session
    ).exclude(
        answer_text__isnull=True
    ).exclude(
        answer_text__exact=""
    ).count()

    progress_percentage = round(
        (
            answered_count
            / total_questions
        ) * 100
    )

    previous_position = (
        current_position - 1
        if current_position > 1
        else None
    )

    next_position = (
        current_position + 1
        if current_position < total_questions
        else None
    )

    context = {
        "session": session,
        "questions": questions,
        "current_question": current_question,
        "current_position": current_position,
        "previous_position": previous_position,
        "next_position": next_position,
        "total_questions": total_questions,
        "answered_count": answered_count,
        "progress_percentage": progress_percentage,
        "is_last_question": (
            current_position == total_questions
        ),
        "completed": (
            session.status == "COMPLETED"
        ),
        "form": form,
    }

    return render(
        request,
        "career_app/interview_session.html",
        context,
    )
@login_required
@transaction.atomic
def complete_interview(request, session_id):
    """
    Complete and evaluate an interview.

    The user must answer every question before the interview can
    be completed.
    """

    session = get_object_or_404(
        InterviewSession.objects.select_related(
            'user',
            'job_role',
            'project',
        ),
        id=session_id,
        user=request.user
    )

    questions = session.questions.all().order_by(
        'display_order',
        'id'
    )

    if not questions.exists():
        messages.error(
            request,
            'This interview does not contain any questions.'
        )

        return redirect(
            'interview_setup'
        )

    answered_question_ids = set(
        InterviewAnswer.objects.filter(
            question__session=session
        ).values_list(
            'question_id',
            flat=True
        )
    )

    unanswered_questions = questions.exclude(
        id__in=answered_question_ids
    )

    if unanswered_questions.exists():
        first_unanswered = unanswered_questions.first()

        messages.warning(
            request,
            'Please answer all questions before completing the interview.'
        )

        interview_url = reverse(
            'interview_session',
            args=[session.id]
        )

        return redirect(
            f'{interview_url}?question={first_unanswered.display_order}'
        )

    empty_answers = InterviewAnswer.objects.filter(
        question__session=session,
        answer_text__regex=r'^\s*$'
    )

    if empty_answers.exists():
        first_empty_answer = empty_answers.select_related(
            'question'
        ).order_by(
            'question__display_order'
        ).first()

        messages.warning(
            request,
            'Please provide an answer for every interview question.'
        )

        interview_url = reverse(
            'interview_session',
            args=[session.id]
        )

        return redirect(
            f'{interview_url}'
            f'?question={first_empty_answer.question.display_order}'
        )

    try:
        evaluate_interview_session(session)

    except Exception as error:
        transaction.set_rollback(True)

        messages.error(
            request,
            f'Interview evaluation failed: {error}'
        )

        return redirect(
            'interview_session',
            session_id=session.id
        )

    messages.success(
        request,
        'Interview completed and evaluated successfully.'
    )

    return redirect(
        'interview_results',
        session_id=session.id
    )


@login_required
def interview_results(request, session_id):
    """
    Display the completed interview evaluation.
    """

    session = get_object_or_404(
        InterviewSession.objects.select_related(
            'user',
            'job_role',
            'project',
            'readiness_assessment',
            'bottleneck',
        ),
        id=session_id,
        user=request.user
    )

    answers = InterviewAnswer.objects.filter(
        question__session=session
    ).select_related(
        'question',
        'question__expected_skill',
        'question__expected_tool',
        'question__competency_group',
    ).order_by(
        'question__display_order',
        'question__id'
    )

    answered_count = answers.count()
    evaluated_count = answers.filter(
        evaluated_at__isnull=False
    ).count()

    context = {
        'session': session,
        'answers': answers,
        'answered_count': answered_count,
        'evaluated_count': evaluated_count,
    }

    return render(
        request,
        'career_app/interview_results.html',
        context
    )

@login_required
def choose_interview_evaluation(request, session_id):
    """
    Allow the user to choose rule-based or AI-powered evaluation.
    """

    session = get_object_or_404(
        InterviewSession.objects.select_related(
            'job_role',
            'project',
        ),
        id=session_id,
        user=request.user,
    )

    total_questions = session.questions.count()

    answered_count = InterviewAnswer.objects.filter(
        question__session=session
    ).exclude(
        answer_text__isnull=True
    ).exclude(
        answer_text__exact=''
    ).count()

    if total_questions == 0:
        messages.error(
            request,
            'This interview does not contain any questions.'
        )
        return redirect('interview_setup')

    if answered_count < total_questions:
        messages.warning(
            request,
            'Please answer every question before choosing an evaluation method.'
        )

        first_unanswered = session.questions.exclude(
            answer__isnull=False
        ).order_by(
            'display_order',
            'id'
        ).first()

        position = (
            first_unanswered.display_order
            if first_unanswered
            else 1
        )

        return redirect(
            f"{reverse('interview_session', args=[session.id])}"
            f"?question={position}"
        )

    if request.method == 'POST':
        evaluation_method = request.POST.get(
            'evaluation_method'
        )

        if evaluation_method not in [
            'rule',
            'ai',
        ]:
            messages.error(
                request,
                'Please select a valid evaluation method.'
            )

            return redirect(
                'choose_interview_evaluation',
                session_id=session.id,
            )

        try:
            evaluate_interview_session(
                session,
                method=evaluation_method,
            )

        except Exception as error:
            messages.error(
                request,
                f'Interview evaluation failed: {error}'
            )

            return redirect(
                'choose_interview_evaluation',
                session_id=session.id,
            )

        messages.success(
            request,
            'Interview evaluated successfully.'
        )

        return redirect(
            'interview_results',
            session_id=session.id,
        )

    context = {
        'session': session,
        'total_questions': total_questions,
        'answered_count': answered_count,
    }

    return render(
        request,
        'career_app/choose_interview_evaluation.html',
        context,
    )

@login_required
def interview_history(request):
    """
    Display all interview sessions created by the logged-in user.
    """

    sessions = InterviewSession.objects.filter(
        user=request.user
    ).select_related(
        'job_role',
        'project',
    ).annotate(
        answer_count=models.Count(
            'questions__answer',
            distinct=True
        ),
        question_count=models.Count(
            'questions',
            distinct=True
        ),
    ).order_by(
        '-created_at'
    )

    return render(
        request,
        'career_app/interview_history.html',
        {
            'sessions': sessions,
        }
    )